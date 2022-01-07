from fhir.models import EncounterModel
import os
import xml.etree.ElementTree as ET
import requests
import openpyxl as xl
from datetime import datetime

ns = {"d": "http://hl7.org/fhir"}

fhir_server = os.getenv('FHIR_SERVER',"http://34.101.85.15:8080/fhir")


def identifier_type(
    resource, system, value, use="usual", _type=None, period=None, assigner=None
):
    ET.SubElement(resource, 'use value="{}"'.format(use))
    if _type:
        type_res = ET.SubElement(resource, "type")
        codeable_concept(type_res, _type)
    ET.SubElement(resource, 'system value="{}"'.format(system))
    ET.SubElement(resource, 'value value="{:0>5}"'.format(value))
    if period:
        period_res = ET.SubElement(resource, "period")
        period_type(period_res, period)
    if assigner:
        assigner_res = ET.SubElement(resource, "assigner")
        reference_type(
            assigner_res, assigner.reference, assigner._type, assigner_res.identifier
        )


def period_type(resource, start=None, end=None):
    if start:
        ET.SubElement(resource, 'start value="{}"'.format(start))
    if end:
        ET.SubElement(resource, 'end value="{}"'.format(end))


def reference_type(resource, reference=None, _type=None, identifier=None, display=None):
    if reference:
        ET.SubElement(resource, 'reference value="{}"'.format(reference))
    if _type:
        ET.SubElement(resource, 'type value="{}"'.format(_type))
    if identifier:
        _identifier = ET.SubElement(resource, "identifier")
        identifier_type(
            _identifier,
            identifier.get("system"),
            identifier.get("value"),
            identifier.get("use"),
            identifier.get("type"),
            identifier.get("period"),
            identifier("assigner"),
        )
    if display:
        ET.SubElement(resource, 'display value="{}"'.format(display))


def name_type(resource, name, text=None, use=None):
    fam_name = name.split(" ")[0]
    given_name = " ".join(name.split(" ")[1:])
    if not use:
        use = "official"
    ET.SubElement(resource, 'use value="{}"'.format(use))
    if text:
        ET.SubElement(resource, 'text value="{}"'.format(text))
    ET.SubElement(resource, 'family value="{}"'.format(fam_name))
    ET.SubElement(resource, 'given value="{}"'.format(given_name))


def address_type(
    resource, address, postalCode=None, country=None, use=None, type_=None
):
    city = address.split(", ")[-1]
    district = address.split(", ")[-2]
    line = " ".join(address.split(", ")[:-2])
    if not use:
        use = "home"
    if not type_:
        type_ = "both"
    if not postalCode:
        postalCode = "70000"
    if not country:
        country = "VietNam"
    ET.SubElement(resource, 'use value="{}"'.format(use))
    ET.SubElement(resource, 'type value="{}"'.format(type_))
    ET.SubElement(resource, 'line value="{}"'.format(line))
    ET.SubElement(resource, 'city value="{}"'.format(city))
    ET.SubElement(resource, 'district value="{}"'.format(district))
    ET.SubElement(resource, 'postalCode value="{}"'.format(postalCode))
    ET.SubElement(resource, 'country value="{}"'.format(country))


def coding_type(
    resource, system, code, display=None, userSelected="false", version="4.0.1"
):
    ET.SubElement(resource, 'system value="{}"'.format(system))
    ET.SubElement(resource, 'version value="{}"'.format(version))
    ET.SubElement(resource, 'code value="{}"'.format(code))
    if display:
        ET.SubElement(resource, 'display value="{}"'.format(display))
    ET.SubElement(resource, 'userSelected value="{}"'.format(userSelected))


def contactpoint_type(resource, system, value, use=None, rank=None, period=None):
    ET.SubElement(resource, 'system value="{}"'.format(system))
    ET.SubElement(resource, 'value value="{}"'.format(value))
    if not use:
        use = "mobile"
    ET.SubElement(resource, 'use value="{}"'.format(use))
    if rank:
        ET.SubElement(resource, 'rank value="{}"'.format(rank))
    if period:
        period_type(resource, period)


def codeable_concept(resource, codes=None, text=None):
    if codes:
        for i in range(len(codes)):
            coding = ET.SubElement(resource, "coding")
            coding_type(
                coding,
                codes[i].get("system"),
                codes[i].get("code"),
                codes[i].get("display"),
                version=codes[i].get("version"),
            )
    if text:
        ET.SubElement(resource, 'text value="{}"'.format(text))


def quantity_type(resource, value, unit, system=None, code=None, comparator=None):
    ET.SubElement(resource, 'value value="{}"'.format(value))
    if comparator:
        ET.SubElement(resource, 'comparator value="{}"'.format(comparator))
    ET.SubElement(resource, 'unit value="{}"'.format(unit))
    if system:
        ET.SubElement(resource, 'system value="{}"'.format(system))
    if code:
        ET.SubElement(resource, 'code value="{}"'.format(code))


def money_type(resource, value, currency):
    ET.SubElement(resource, 'value value="{}"'.format(value))
    ET.SubElement(resource, 'currency value="{}"'.format(currency))


def range_type(resource, low_limit, high_limit):
    low = ET.SubElement(resource, "low")
    quantity_type(
        low,
        low_limit.value,
        low_limit.comparator,
        low_limit.unit,
        low_limit.system,
        low_limit.code,
    )
    high = ET.SubElement(resource, "high")
    quantity_type(
        high,
        high_limit.value,
        high_limit.comparator,
        high_limit.unit,
        high_limit.system,
        high_limit.code,
    )


def ratio_type(resource, numerator_value, denominator_value):
    numerator = ET.SubElement(resource, "numerator")
    quantity_type(
        numerator,
        numerator_value.value,
        numerator_value.comparator,
        numerator_value.unit,
        numerator_value.system,
        numerator_value.code,
    )
    denominator = ET.SubElement(resource, "denominator")
    quantity_type(
        denominator,
        denominator_value.value,
        denominator_value.comparator,
        denominator_value.unit,
        denominator_value.system,
        denominator_value.code,
    )


def duration_type(resource, value, unit, system, code):
    ET.SubElement(resource, 'value value="{}"'.format(value))
    ET.SubElement(resource, 'unit value="{}"'.format(unit))
    ET.SubElement(resource, 'system value="{}"'.format(system))
    ET.SubElement(resource, 'code value="{}"'.format(code))


def annotation_type(resource, text, author=None, time=None):
    if author:
        ET.SubElement(resource, "author")
    if time:
        ET.SubElement(resource, "time")
        time.set("value", time)
    textobject = ET.SubElement(resource, "text")
    textobject.set("value", text)


def timing_type(
    resource,
    duration_value,
    frequency_value,
    period_value,
    when_value,
    offset_value=None,
    event_value=None,
):
    if event_value:
        event = ET.SubElement(resource, "event")
        event.set("value", event_value)
    repeat = ET.SubElement(resource, "repeat")
    get_duration = duration_value.split(" ")
    duration = ET.SubElement(repeat, "duration")
    if "-" in get_duration[0]:
        duration.set("value", get_duration[0].split("-")[0])
        durationMax = ET.SubElement(repeat, "durationMax")
        durationMax.set("value", get_duration[0].split("-")[1])
    else:
        duration.set("value", get_duration[0])
    durationUnit = ET.SubElement(repeat, "durationUnit")
    durationUnit.set("value", get_duration[1])
    get_frequency = frequency_value.split(" ")
    frequency = ET.SubElement(repeat, "frequency")
    if "-" in get_frequency[0]:
        frequency.set("value", get_frequency[0].split("-")[0])
        frequencyMax = ET.SubElement(repeat, "frequencyMax")
        frequencyMax.set("value", get_frequency[0].split("-")[1])
    else:
        frequency.set("value", get_frequency[0])
    get_period = period_value.split(" ")
    period = ET.SubElement(repeat, "period")
    if "-" in get_period[0]:
        period.set("value", get_period[0].split("-")[0])
        periodMax = ET.SubElement(repeat, "periodMax")
        periodMax.set("value", get_period[0].split("-")[1])
    else:
        period.set("value", get_period[0])
    periodUnit = ET.SubElement(repeat, "periodUnit")
    periodUnit.set("value", get_period[1])
    when = ET.SubElement(repeat, "when")
    when.set("value", when_value)
    if offset_value:
        offset = ET.SubElement(repeat, "offset")
        offset.set("value", offset_value)


def get_observation(encounter_id):
    get_observation = requests.get(
        fhir_server + "/Observation?encounter=" + str(encounter_id),
        headers={"Content-type": "application/xml"},
    )
    if (
        get_observation.status_code == 200
        and "entry" in get_observation.content.decode("utf-8")
    ):
        get_root = ET.fromstring(get_observation.content.decode("utf-8"))
        observation = []
        for entry in get_root.findall("d:entry", ns):
            resource = entry.find("d:resource", ns)
            observation_resource = resource.find("d:Observation", ns)
            code = observation_resource.find("d:code", ns)
            coding = code.find("d:coding", ns)
            display = coding.find("d:display", ns).attrib["value"]
            quantity = observation_resource.find("d:valueQuantity", ns)
            value = (
                quantity.find("d:value", ns).attrib["value"]
                + quantity.find("d:unit", ns).attrib["value"]
            )
            observation.append({"display": display, "value": value})
        return observation
    else:
        return None


def getdatetime(datetime_str):
    get_datetime = datetime_str.split("+")[0]
    if "." in get_datetime:
        get_datetime = get_datetime.split(".")[0]
    # print(get_datetime)
    datetime_obj = datetime.strptime(get_datetime, "%Y-%m-%dT%H:%M:%S")
    return datetime.strftime(datetime_obj, "%Y-%m-%d %H:%M:%S")


def get_date(date_str):
    date = datetime.strptime(date_str, "%Y-%m-%d")
    return datetime.strftime(date, "%d-%m-%Y")


# def get_encounter(encounter_id):
#     encounter = {}
#     get_encounter = requests.get(fhir_server + "/Encounter/" + str(
#         encounter_id), headers={'Content-type': 'application/xml'})
#     print(get_encounter.status_code)
#     print(get_encounter.content)
#     if get_encounter.status_code == 200:
#         get_root = ET.fromstring(get_encounter.content.decode('utf-8'))
#         encounter['id'] = get_root.find('d:id', ns).attrib['value']
#         period = get_root.find('d:period', ns)
#         encounter['start_date'] = period.find('d:start', ns).attrib['value']
#         encounter['id'] = encounter_id
#         return encounter
#     else:
#         return None


def get_patient_upload_data(excel_file):
    wb = xl.load_workbook(excel_file)
    sh = wb["Sheet1"]
    max_row = sh.max_row
    data = {"Patient": {}, "Encounter": {}, "Observation": [], "Condition": []}
    for i in range(1, max_row + 1):
        cell = sh.cell(row=i, column=5)
        if cell.value:
            tag = sh.cell(row=i, column=4)
            if "encounter" in tag.value:
                data["Encounter"][tag.value] = cell.value
            elif "conditions" in tag.value:
                if "admission" in tag.value:
                    if tag.value.split(".")[1] == "patient":
                        data["Condition"]["admission_by_patient"] = []
                        for condition in cell.value.split(";"):

                            data["Condition"]["admission_by_patient"].append()
                        pass
                    else:
                        data["Condition"]["admission_by_doctor"] = []
                        pass
                    pass
                elif "resolved" in tag.value:
                    data["Condition"]["resolved_conditions"] = []
                    pass
                elif "discharge" in tag.value:
                    data["Condition"]["discharge_conditions"] = []
                    pass
                elif "comorbidity" in tag.value:
                    data["Condition"]["comorbidity_conditions"] = []
                    pass
            elif "observations" in tag.value:
                data["Observation"].append({tag.value.split(".")[1]: cell.value})
            elif "diagnostic" in tag.value:
                pass
            elif "service" in tag.value:
                pass
            else:
                data["Patient"][tag.value] = cell.value

    pass


def get_patient_upload_data(excel_file):
    wb = xl.load_workbook(excel_file)
    sh = wb["Sheet1"]
    m_row = sh.max_row
    last_line = 0
    data = {"Patient": {}, "Encounter": {}, "Observation": []}
    for i in range(1, m_row + 1):
        cell = sh.cell(row=i, column=4)
        if cell.value:
            tag = sh.cell(row=i, column=2)
            if tag.value == "HO_TEN":
                data["Patient"]["name"] = cell.value
            elif tag.value == "MA_DKBD":
                data["Patient"]["identifier"] = cell.value
            elif tag.value == "NGAY_SINH":
                data["Patient"]["birthDate"] = cell.value
            elif tag.value == "GIOI_TINH":
                data["Patient"]["gender"] = cell.value
            elif tag.value == "DIA_CHI":
                data["Patient"]["address"] = [{"address": cell.value}]
            elif tag.value == "NGAY_VAO":
                data["Encounter"]["period"] = {"start": cell.value}
                data["Encounter"]["start_date"] = cell.value
            elif tag.value == "NGAY_RA":
                if data["Encounter"].get("period"):
                    data["Encounter"]["period"]["stop"] = cell.value
                else:
                    data["Encounter"]["period"] = {"stop": cell.value}
            elif tag.value == "MA_LOAI_KCB":
                data["Encounter"]["class"] = cell.value
            elif tag.value == "MA_KHOA":
                data["Encounter"]["location"] = cell.value
            elif tag.value == "SO_NGAY_DTRI":
                data["Encounter"]["length"] = cell.value
            elif not tag.value:
                tag_2 = sh.cell(row=i, column=3)
                tag_2_content = tag_2.value.split(".")
                if tag_2_content[0] == "Patient":
                    if data["Patient"][tag_2_content[1]]:
                        data["Patient"][tag_2_content[1]].append(
                            {tag_2_content[1]: cell.value}
                        )
                        if len(tag_2_content) > 2:
                            for i in range(2, len(tag_2_content)):
                                data["Patient"][tag_2_content[1]][-1][
                                    tag_2_content[i].split("=")[0]
                                ] = tag_2_content[i].split("=")[1]
                elif tag_2_content[0] == "Encounter":
                    if data["Encounter"][tag_2_content[1]]:
                        data["Encounter"][tag_2_content[1]].append(
                            {tag_2_content[1]: cell.value}
                        )
                        if len(tag_2_content) > 2:
                            for i in range(2, len(tag_2_content)):
                                data["Encounter"][tag_2_content[1]][-1][
                                    tag_2_content[i].split("=")[0]
                                ] = tag_2_content[i].split("=")[1]
                elif tag_2_content[0] == "Observation":
                    _observation = {}
                    for i in range(1, len(tag_2_content)):
                        _observation[tag_2_content[i].split("=")[0]] = tag_2_content[
                            i
                        ].split("=")[1]
                    _observation["valueQuantity"] = cell.value
                    data["Observation"].append(_observation)
    # print(data)
    return data


def create_patient_resource(patient):
    root = ET.Element("Patient")
    tree = ET.ElementTree(root)
    root.set("xmlns", "http://hl7.org/fhir")
    if patient.get("id"):
        id_ = ET.SubElement(root, "id")
        id_.set("value", patient["id"])
    if patient.get("identifier"):
        identifier_resource = ET.SubElement(root, "identifier")
        identifier_type(
            identifier_resource,
            "urn:trinhcongminh",
            patient["identifier"],
            "usual",
            [{"system": "http://terminology.hl7.org/CodeSystem/v2-0203", "code": "SB"}],
        )
    if patient.get("name"):
        name = ET.SubElement(root, "name")
        name_type(name, patient["name"])
    if patient.get("telecom"):
        telecom = ET.SubElement(root, "telecom")
        contactpoint_type(telecom, "phone", patient["telecom"])
    if patient.get("gender"):
        gender = ET.SubElement(root, "gender")
        if patient["gender"] == "Nam":
            code = "male"
        elif patient["gender"] == "Nữ":
            code = "female"
        gender.set("value", code)
    if patient.get("birthdate"):
        birthDate = ET.SubElement(root, "birthDate")
        birthDate.set("value", patient["birthdate"])
    if patient.get("home_address"):
        address = ET.SubElement(root, "address")
        address_type(address, patient["home_address"], use="home")
    if patient.get("work_address"):
        address = ET.SubElement(root, "address")
        address_type(address, patient["work_address"], use="work")
    contact = ET.SubElement(root, "contact")
    if patient.get("contact_relationship"):
        contact_relationship = ET.SubElement(contact, "relationship")
        codes = [
            {
                "system": "http://terminology.hl7.org/CodeSystem/v2-0131",
                "code": patient["contact_relationship"],
                "version": "4.0.1",
            }
        ]
        # codes = [
        #     {'system': 'http://hl7.org/fhir/reaction-event-severity', 'code': allergy['reaction']['severity']}
        # ]
        codeable_concept(
            contact_relationship, codes, text=patient["contact_relationship"]
        )
    if patient.get("contact_name"):
        name = ET.SubElement(contact, "name")
        name_type(name, patient["contact_name"])
    if patient.get("contact_telecom"):
        telecom = ET.SubElement(contact, "telecom")
        contactpoint_type(telecom, "phone", patient["contact_telecom"])
    if patient.get("contact_address"):
        address = ET.SubElement(contact, "address")
        address_type(address, patient["contact_address"])
    if patient.get("contact_gender"):
        gender = ET.SubElement(contact, "gender")
        if patient["contact_gender"] == "Nam":
            gender.set("value", "male")
        else:
            gender.set("value", "female")
    return ET.tostring(
        root,
        encoding="us-ascii",
        method="xml",
        xml_declaration=None,
        default_namespace=None,
        short_empty_elements=True,
    )


def create_practitioner_resource(practitioner):
    root = ET.Element("Practitioner")
    tree = ET.ElementTree(root)
    root.set("xmlns", "http://hl7.org/fhir")
    if practitioner.get("id"):
        id_ = ET.SubElement(root, "id")
        id_.set("value", practitioner["id"])
    if practitioner.get("identifier"):
        identifier_resource = ET.SubElement(root, "identifier")
        identifier_type(
            identifier_resource,
            "urn:trinhcongminh",
            practitioner["identifier"],
            "usual",
            [{"system": "http://terminology.hl7.org/CodeSystem/v2-0203", "code": "EI"}],
        )
    if practitioner.get("name"):
        name = ET.SubElement(root, "name")
        name_type(name, practitioner["name"])
    if practitioner.get("telecom"):
        telecom = ET.SubElement(root, "telecom")
        contactpoint_type(telecom, "phone", practitioner["telecom"])
    if practitioner.get("gender"):
        gender = ET.SubElement(root, "gender")
        if practitioner["gender"] == "Nam":
            code = "male"
        elif practitioner["gender"] == "Nữ":
            code = "female"
        gender.set("value", code)
    if practitioner.get("birthdate"):
        birthdate = ET.SubElement(root, "birthDate")
        birthdate.set("value", practitioner["birthdate"])
    if practitioner.get("home_address"):
        address = ET.SubElement(root, "address")
        address_type(address, practitioner["home_address"], use="home")
    if practitioner.get("qualification"):
        qualification = ET.SubElement(root, "qualification")
        qualification_code = ET.SubElement(qualification, "code")
        codeable_concept(qualification_code, text=practitioner["qualification"])
    return ET.tostring(
        root,
        encoding="us-ascii",
        method="xml",
        xml_declaration=None,
        default_namespace=None,
        short_empty_elements=True,
    )


def create_encounter_resource(
    encounter, patient_id, patient_name, practitioner_id, practitioner_name
):
    root = ET.Element("Encounter")
    tree = ET.ElementTree(root)
    root.set("xmlns", "http://hl7.org/fhir")
    # identifier = ET.SubElement(root, 'identifier')
    # dttype.identifier_type(identifier, )
    if encounter.get("id"):
        id_ = ET.SubElement(root, "id")
        id_.set("value", encounter["id"])
    if encounter.get("identifier"):
        identifier_resource = ET.SubElement(root, "identifier")
        identifier_type(
            identifier_resource,
            "urn:trinhcongminh",
            encounter["identifier"],
            "usual",
            [{"system": "http://terminology.hl7.org/CodeSystem/v2-0203", "code": "MR"}],
        )
    if not encounter.get("status"):
        status = ET.SubElement(root, "status")
        status.set("value", "in-progress")
    else:
        status = ET.SubElement(root, "status")
        status.set("value", encounter["status"])
    if encounter.get("class"):
        _class = ET.SubElement(root, "class")
        coding_type(
            _class,
            "http://terminology.hl7.org/CodeSystem/v3-ActCode",
            encounter["class"],
        )
    if encounter.get("type"):
        _type = ET.SubElement(root, "type")
        codeable_concept(_type, text=encounter["type"])
    if encounter.get("service_type"):
        service_type = ET.SubElement(root, "serviceType")
        codeable_concept(service_type, text=encounter["service_type"])
    if encounter.get("priority"):
        priority = ET.SubElement(root, "priority")
        priority_code = [
            {
                "system": "http://terminology.hl7.org/CodeSystem/v3-ActPriority",
                "code": encounter["priority"]["code"],
                "version": "2018-08-12",
            }
        ]
        codeable_concept(priority, priority_code, text=encounter["priority"]["display"])
    subject = ET.SubElement(root, "subject")
    reference_type(subject, "Patient/" + patient_id, "Patient", display=patient_name)
    participant = ET.SubElement(root, "participant")
    participant_type = ET.SubElement(participant, "type")
    participant_type_codes = [
        {
            "system": "http://terminology.hl7.org/CodeSystem/v3-ParticipationType",
            "code": "DIS",
            "version": "2018-08-12",
        }
    ]
    codeable_concept(participant_type, participant_type_codes)
    individual = ET.SubElement(participant, "individual")
    reference_type(
        individual,
        "Practitioner/" + practitioner_id,
        "Practitioner",
        display=practitioner_name,
    )
    if encounter.get("period"):
        period = ET.SubElement(root, "period")
        period_type(
            period, encounter["period"].get("start"), encounter["period"].get("end")
        )
    if encounter.get("length"):
        length = ET.SubElement(root, "length")
        duration_type(
            length, encounter["length"], "days", "http://unitsofmeasure.org", "d"
        )
    if encounter.get("reason_code"):
        reason = ET.SubElement(root, "reasonCode")
        codeable_concept(reason, text=encounter["reason_code"])
    if encounter.get("location"):
        location = ET.SubElement(root, "location")
        location_reference = ET.SubElement(location, "location")
        reference_type(location_reference, display=encounter["location"])
    if encounter.get("serviceProvider"):
        serviceProvider = ET.SubElement(root, "serviceProvider")
    return ET.tostring(
        root,
        encoding="us-ascii",
        method="xml",
        xml_declaration=None,
        default_namespace=None,
        short_empty_elements=True,
    )


def create_condition_resource(condition, patient_id, patient_name, encounter_id):
    root = ET.Element("Condition")
    tree = ET.ElementTree(root)
    root.set("xmlns", "http://hl7.org/fhir")
    if condition.get("id"):
        id_ = ET.SubElement(root, "id")
        id_.set("value", condition["id"])
    if condition.get("identifier"):
        identifier = ET.SubElement(root, "identifier")
        identifier_type(
            identifier, "urn:trinhcongminh", condition["identifier"], "usual"
        )
    if condition.get("clinical_status"):
        clinical = ET.SubElement(root, "clinicalStatus")
        codeable_concept(
            clinical,
            [
                {
                    "system": "http://terminology.hl7.org/CodeSystem/condition-clinical",
                    "code": condition["clinical_status"],
                    "version": "4.0.1",
                }
            ],
            text=condition["clinical_status"],
        )
    if condition.get("verification_status"):
        verify = ET.SubElement(root, "verificationStatus")
        codeable_concept(
            verify,
            [
                {
                    "system": "http://terminology.hl7.org/CodeSystem/condition-ver-status",
                    "code": "confirmed",
                    "version": "4.0.1",
                }
            ],
            text="confirmed",
        )
    if condition.get("category"):
        category = ET.SubElement(root, "category")
        codeable_concept(category, text=condition["category"])
    if condition.get("severity"):
        severity = ET.SubElement(root, "severity")
        codeable_concept(
            severity,
            [
                {
                    "system": "http://snomed.info/sct",
                    "code": condition["severity"]["code"],
                    "version": "4.0.1",
                }
            ],
            text=condition["severity"]["display"],
        )
    if condition.get("code"):
        if condition.get("display_code"):
            code = ET.SubElement(root, "code")
            codes = [
                {"system": "http://hl7.org/fhir/sid/icd-10", "code": condition["code"]}
            ]
            codeable_concept(
                code,
                codes=codes,
                text=condition["code"] + ": " + condition["display_code"],
            )
        else:
            code = ET.SubElement(root, "code")
            codeable_concept(code, text=condition["code"])
    subject = ET.SubElement(root, "subject")
    reference_type(subject, "Patient/" + patient_id, "Patient", display=patient_name)
    encounter = ET.SubElement(root, "encounter")
    reference_type(encounter, "Encounter/" + encounter_id, "Encounter")
    if condition.get("onset"):
        onset = ET.SubElement(root, "onsetDateTime")
        onset.set("value", condition["onset"])
    if condition.get("abatement"):
        abatement = ET.SubElement(root, "abatementDateTime")
        abatement.set("value", condition["abatement"])
    if condition.get("asserter"):
        asserter = ET.SubElement(root, "asserter")
        reference_type(
            asserter,
            condition["asserter"]["type"] + "/" + condition["asserter"]["id"],
            condition["asserter"]["type"],
            display=condition["asserter"]["name"],
        )
    if condition.get("note"):
        note = ET.SubElement(root, "note")
        annotation_type(note, text=condition["note"])
    return ET.tostring(
        root,
        encoding="us-ascii",
        method="xml",
        xml_declaration=None,
        default_namespace=None,
        short_empty_elements=True,
    )


def create_service_resource(service, patient_id, patient_name, encounter_id):
    root = ET.Element("ServiceRequest")
    tree = ET.ElementTree(root)
    root.set("xmlns", "http://hl7.org/fhir")
    if service.get("id"):
        id_ = ET.SubElement(root, "id")
        id_.set("value", service["id"])
    if service.get("identifier"):
        identifier = ET.SubElement(root, "identifier")
        identifier_type(identifier, "urn:trinhcongminh", service["identifier"], "usual")
    if service.get("status"):
        status = ET.SubElement(root, "status")
        status.set("value", service["status"])
    if service.get("intent"):
        intent = ET.SubElement(root, "intent")
        intent.set("value", service["intent"])
    if service.get("category"):
        category = ET.SubElement(root, "category")
        codeable_concept(category, text=service["category"])
    if service.get("code"):
        code = ET.SubElement(root, "code")
        codeable_concept(code, text=service["code"])
    subject = ET.SubElement(root, "subject")
    reference_type(subject, "Patient/" + patient_id, "Patient", display=patient_name)
    encounter = ET.SubElement(root, "encounter")
    reference_type(encounter, "Encounter/" + encounter_id, "Encounter")
    if service.get("occurrence"):
        occurrence = ET.SubElement(root, "occurrenceDateTime")
        occurrence.set("value", service["occurrence"])
    if service.get("authoredOn"):
        authoredOn = ET.SubElement(root, "authoredOn")
        authoredOn.set("value", service["authoredOn"])
    if service.get("requester"):
        requester = ET.SubElement(root, "requester")
        reference_type(
            requester,
            "Practitioner/" + service["requester"]["id"],
            "Practitioner",
            display=service["requester"]["name"],
        )
    if service.get("performer"):
        performer = ET.SubElement(root, "performer")
        reference_type(
            performer,
            "Practitioner/" + service["performer"]["id"],
            "Practitioner",
            display=service["performer"]["name"],
        )
    if service.get("reason_code"):
        reason_code = ET.SubElement(root, "reasonCode")
        codeable_concept(reason_code, text=service["reason_code"])
    if service.get("note"):
        note = ET.SubElement(root, "note")
        annotation_type(note, text=service["note"])
    return ET.tostring(
        root,
        encoding="us-ascii",
        method="xml",
        xml_declaration=None,
        default_namespace=None,
        short_empty_elements=True,
    )


def create_observation_resource(
    observation, patient_id, patient_name, encounter_id, service_id=None
):
    root = ET.Element("Observation")
    tree = ET.ElementTree(root)
    root.set("xmlns", "http://hl7.org/fhir")
    if observation.get("id"):
        id_ = ET.SubElement(root, "id")
        id_.set("value", observation["id"])
    if observation.get("identifier"):
        identifier = ET.SubElement(root, "identifier")
        identifier_type(
            identifier, "urn:trinhcongminh", observation["identifier"], "usual"
        )
    if service_id:
        based_on = ET.SubElement(root, "basedOn")
        reference_type(based_on, "ServiceRequest/" + service_id, "ServiceRequest")
    if observation.get("status"):
        status = ET.SubElement(root, "status")
        status.set("value", observation["status"])
    if observation.get("category"):
        category = ET.SubElement(root, "category")
        codeable_concept(
            category,
            [
                {
                    "system": "http://terminology.hl7.org/CodeSystem/observation-category",
                    "code": observation["category"],
                    "version": "4.0.1",
                }
            ],
            text=observation["category"],
        )
    if observation.get("code"):
        code = ET.SubElement(root, "code")
        codeable_concept(code, text=observation["code"])
    subject = ET.SubElement(root, "subject")
    reference_type(subject, "Patient/" + patient_id, "Patient", display=patient_name)
    encounter = ET.SubElement(root, "encounter")
    reference_type(encounter, "Encounter/" + encounter_id, "Encounter")
    if observation.get("effective"):
        effectiveDateTime = ET.SubElement(root, "effectiveDateTime")
        effectiveDateTime.set("value", observation["effective"])
    if observation.get("performer"):
        performer = ET.SubElement(root, "performer")
        reference_type(
            performer,
            "Practitioner/" + observation["performer"]["id"],
            "Practitioner",
            display=observation["performer"]["name"],
        )
    if observation.get("value_quantity"):
        valueQuantity = ET.SubElement(root, "valueQuantity")
        quantity_type(
            valueQuantity, observation["value_quantity"], observation["value_unit"]
        )
    if observation.get("note"):
        note = ET.SubElement(root, "note")
        annotation_type(note, text=observation["note"])
    if observation.get("reference_range"):
        reference_range = ET.SubElement(root, "referenceRange")
        reference_range_text = ET.SubElement(reference_range, "text")
        reference_range_text.set("value", observation["reference_range"])
    return ET.tostring(
        root,
        encoding="us-ascii",
        method="xml",
        xml_declaration=None,
        default_namespace=None,
        short_empty_elements=True,
    )


def create_procedure_resource(
    procedure, patient_id, patient_name, encounter_id, service_id
):
    root = ET.Element("Procedure")
    tree = ET.ElementTree(root)
    root.set("xmlns", "http://hl7.org/fhir")
    if procedure.get("id"):
        id_ = ET.SubElement(root, "id")
        id_.set("value", procedure["id"])
    if procedure.get("identifier"):
        identifier = ET.SubElement(root, "identifier")
        identifier_type(
            identifier, "urn:trinhcongminh", procedure["identifier"], "usual"
        )
    if service_id:
        based_on = ET.SubElement(root, "basedOn")
        reference_type(based_on, "ServiceRequest/" + service_id, "ServiceRequest")
    if procedure.get("status"):
        status = ET.SubElement(root, "status")
        status.set("value", procedure["status"])
    if procedure.get("category"):
        category = ET.SubElement(root, "category")
        codeable_concept(category, text=procedure["category"])
    if procedure.get("code"):
        code = ET.SubElement(root, "code")
        codeable_concept(code, text=procedure["code"])
    subject = ET.SubElement(root, "subject")
    reference_type(subject, "Patient/" + patient_id, "Patient", display=patient_name)
    encounter = ET.SubElement(root, "encounter")
    reference_type(encounter, "Encounter/" + encounter_id, "Encounter")
    if procedure.get("performed_datetime"):
        performed_dateTime = ET.SubElement(root, "performedDateTime")
        performed_dateTime.set("value", procedure["performed_datetime"])
    if procedure.get("asserter"):
        asserter = ET.SubElement(root, "asserter")
        reference_type(
            asserter,
            "Practitioner/" + procedure["asserter"]["id"],
            "Practitioner",
            display=procedure["asserter"]["name"],
        )
    if procedure.get("performer"):
        performer = ET.SubElement(root, "performer")
        actor = ET.SubElement(performer, "actor")
        reference_type(
            actor,
            "Practitioner/" + procedure["performer"]["id"],
            "Practitioner",
            display=procedure["performer"]["name"],
        )
    if procedure.get("reason_code"):
        reasonCode = ET.SubElement(root, "reasonCode")
        codeable_concept(reasonCode, text=procedure["reason_code"])
    if procedure.get("outcome"):
        outcome = ET.SubElement(root, "outcome")
        codes = [
            {
                "system": "http://snomed.info/sct",
                "code": procedure["outcome"],
                "version": "4.0.1",
            }
        ]
        codeable_concept(outcome, codes, text=procedure["outcome"])
    if procedure.get("complication"):
        complication = ET.SubElement(root, "complication")
        codeable_concept(complication, text=procedure["complication"])
    if procedure.get("follow_up"):
        followUp = ET.SubElement(root, "followUp")
        codeable_concept(followUp, text=procedure["follow_up"])
    if procedure.get("note"):
        note = ET.SubElement(root, "note")
        annotation_type(note, text=procedure["note"])
    if procedure.get("used"):
        used = ET.SubElement(root, "usedCode")
        codeable_concept(used, text=procedure["used"])
    return ET.tostring(
        root,
        encoding="us-ascii",
        method="xml",
        xml_declaration=None,
        default_namespace=None,
        short_empty_elements=True,
    )


def create_medication_resource(medication, patient_id, patient_name, encounter_id):
    root = ET.Element("MedicationStatement")
    tree = ET.ElementTree(root)
    root.set("xmlns", "http://hl7.org/fhir")
    if medication.get("id"):
        id_ = ET.SubElement(root, "id")
        id_.set("value", medication["id"])
    if medication.get("identifier"):
        identifier = ET.SubElement(root, "identifier")
        identifier_type(
            identifier, "urn:trinhcongminh", medication["identifier"], "usual"
        )
    if medication.get("status"):
        status = ET.SubElement(root, "status")
        status.set("value", medication["status"])
    if medication.get("medication"):
        medicationobject = ET.SubElement(root, "medicationCodeableConcept")
        codeable_concept(medicationobject, text=medication["medication"])
    subject = ET.SubElement(root, "subject")
    reference_type(subject, "Patient/" + patient_id, "Patient", display=patient_name)
    context = ET.SubElement(root, "context")
    reference_type(context, "Encounter/" + encounter_id, "Encounter")
    if medication.get("effective"):
        effective = ET.SubElement(root, "effectiveDateTime")
        effective.set("value", medication["effective"])
        # print(medication['effective'])
    if medication.get("date_asserted"):
        dateAsserted = ET.SubElement(root, "dateAsserted")
        dateAsserted.set("value", medication["date_asserted"])

    if medication.get("reason_code"):
        reasonCode = ET.SubElement(root, "reasonCode")
        codeable_concept(reasonCode, text=medication["reason_code"])
    dosage = ET.SubElement(root, "dosage")
    if medication.get("additional_instruction"):
        additionalInstruction = ET.SubElement(dosage, "additionalInstruction")
        codeable_concept(
            additionalInstruction, text=medication["additional_instruction"]
        )
    if medication.get("patient_instruction"):
        patientInstruction = ET.SubElement(dosage, "patientInstruction")
        patientInstruction.set("value", medication["patient_instruction"])
    timing = ET.SubElement(dosage, "timing")
    if (
        medication.get("frequency")
        and medication.get("period")
        and medication.get("duration")
        and medication.get("when")
    ):
        timing_type(
            timing,
            medication["duration"],
            medication["frequency"],
            medication["period"],
            medication["when"],
            medication.get("offset"),
        )
    if medication.get("route"):
        route = ET.SubElement(dosage, "route")
        codeable_concept(route, text=medication["route"])
    if medication.get("quantity"):
        doseAndRate = ET.SubElement(dosage, "doseAndRate")
        doseQuantity = ET.SubElement(doseAndRate, "doseQuantity")
        dose_value = medication["quantity"].split(" ")[0]
        dose_unit = medication["quantity"].split(" ")[1]
        quantity_type(doseQuantity, dose_value, dose_unit)
    return ET.tostring(
        root,
        encoding="us-ascii",
        method="xml",
        xml_declaration=None,
        default_namespace=None,
        short_empty_elements=True,
    )


def create_diagnostic_report_resource(
    diagnostic_report, patient_id, patient_name, encounter_id, service_id
):
    print(diagnostic_report)
    root = ET.Element("DiagnosticReport")
    tree = ET.ElementTree(root)
    root.set("xmlns", "http://hl7.org/fhir")
    if diagnostic_report.get("id"):
        id_ = ET.SubElement(root, "id")
        id_.set("value", diagnostic_report["id"])
    if diagnostic_report.get("identifier"):
        identifier = ET.SubElement(root, "identifier")
        identifier_type(
            identifier, "urn:trinhcongminh", diagnostic_report["identifier"], "usual"
        )
    based_on = ET.SubElement(root, "basedOn")
    reference_type(based_on, "ServiceRequest/" + service_id, "ServiceRequest")
    if diagnostic_report.get("status"):
        status = ET.SubElement(root, "status")
        status.set("value", diagnostic_report["status"])
    if diagnostic_report.get("category"):
        category = ET.SubElement(root, "category")
        codeable_concept(category, text=diagnostic_report["category"])
    if diagnostic_report.get("code"):
        code = ET.SubElement(root, "code")
        codeable_concept(code, text=diagnostic_report["code"])
    subject = ET.SubElement(root, "subject")
    reference_type(subject, "Patient/" + patient_id, "Patient", display=patient_name)
    encounter = ET.SubElement(root, "encounter")
    reference_type(encounter, "Encounter/" + encounter_id, "Encounter")
    if diagnostic_report.get("effective"):
        effective_datetime = ET.SubElement(root, "effectiveDateTime")
        effective_datetime.set("value", diagnostic_report["effective"])
    if diagnostic_report.get("performer"):
        performer = ET.SubElement(root, "performer")
        reference_type(
            performer,
            "Practitioner/" + diagnostic_report["performer"]["id"],
            "Practitioner",
            display=diagnostic_report["performer"]["name"],
        )
    if diagnostic_report.get("conclusion"):
        conclusion = ET.SubElement(root, "conclusion")
        conclusion.set("value", diagnostic_report["conclusion"])
    return ET.tostring(
        root,
        encoding="us-ascii",
        method="xml",
        xml_declaration=None,
        default_namespace=None,
        short_empty_elements=True,
    )


def create_allergy_resource(allergy, patient_id, patient_name, encounter_id):
    root = ET.Element("AllergyIntolerance")
    tree = ET.ElementTree(root)
    root.set("xmlns", "http://hl7.org/fhir")
    if allergy.get("id"):
        id_ = ET.SubElement(root, "id")
        id_.set("value", allergy["id"])
    if allergy.get("identifier"):
        identifier = ET.SubElement(root, "identifier")
        identifier_type(identifier, "urn:trinhcongminh", allergy["identifier"])
    if allergy.get("clinical_status"):
        clinical_status = ET.SubElement(root, "clinicalStatus")
        codes = [
            {
                "system": "http://terminology.hl7.org/CodeSystem/allergyintolerance-clinical",
                "code": allergy["clinical_status"]["code"],
                "version": "4.0.1",
            }
        ]
        codeable_concept(
            clinical_status, codes, text=allergy["clinical_status"]["display"]
        )
    if allergy.get("verification_status"):
        verification_status = ET.SubElement(root, "verificationStatus")
        codes = [
            {
                "system": "http://terminology.hl7.org/CodeSystem/allergyintolerance-verification",
                "code": allergy["verification_status"],
                "version": "4.0.1",
            }
        ]
        codeable_concept(
            verification_status, codes, text=allergy["verification_status"]
        )
    allergy_type = ET.SubElement(root, "type")
    allergy_type.set("value", "allergy")
    if allergy.get("category"):
        category = ET.SubElement(root, "category")
        # codes = [
        #     {'system': 'http://hl7.org/fhir/allergy-intolerance-category', 'code': allergy['category'], 'version': '4.0.1'}
        # ]
        # codeable_concept(category, codes, text=allergy['category'])
        category.set("value", allergy["category"])
    if allergy.get("criticality"):
        criticality = ET.SubElement(root, "criticality")
        # codes = [
        #     {'system': 'http://hl7.org/fhir/allergy-intolerance-criticality', 'code': allergy['criticality'], 'version': '4.0.1'}
        # ]
        criticality.set("value", allergy["criticality"])
        # codeable_concept(criticality, codes, text=allergy['criticality'])
    if allergy.get("code"):
        code = ET.SubElement(root, "code")
        codeable_concept(code, text=allergy["code"])
    patient = ET.SubElement(root, "patient")
    reference_type(patient, "Patient/" + patient_id, "Patient", display=patient_name)
    encounter = ET.SubElement(root, "encounter")
    reference_type(encounter, "Encounter/" + encounter_id, "Encounter")
    if allergy.get("onset"):
        onset = ET.SubElement(root, "onsetDateTime")
        onset.set("value", allergy["onset"])
    if allergy.get("last_occurrence"):
        last_occurrence = ET.SubElement(root, "lastOccurrence")
        last_occurrence.set("value", allergy["last_occurrence"])
    if allergy.get("reaction"):
        reaction = ET.SubElement(root, "reaction")
        if allergy["reaction"].get("substance"):
            substance = ET.SubElement(reaction, "substance")
            codeable_concept(substance, text=allergy["reaction"]["substance"])
        if allergy["reaction"].get("manifestation"):
            manifestation = ET.SubElement(reaction, "manifestation")
            codeable_concept(manifestation, text=allergy["reaction"]["manifestation"])
        if allergy["reaction"].get("severity"):
            severity = ET.SubElement(reaction, "severity")
            severity.set("value", allergy["reaction"]["severity"])
            # codes = [
            #     {'system': 'http://hl7.org/fhir/reaction-event-severity', 'code': allergy['reaction']['severity']}
            # ]
            # codeable_concept(severity, codes, text=allergy['reaction']['severity'])
    return ET.tostring(
        root,
        encoding="us-ascii",
        method="xml",
        xml_declaration=None,
        default_namespace=None,
        short_empty_elements=True,
    )


def query_patient(patient_identifier, query_type):
    patient = {}
    get_req = requests.get(
        fhir_server + "/Patient?identifier=urn:trinhcongminh|" + patient_identifier,
        headers={"Content-type": "application/xml"},
    )
    if get_req.status_code == 200 and "entry" in get_req.content.decode("utf-8"):
        get_root = ET.fromstring(get_req.content.decode("utf-8"))
        entry = get_root.find("d:entry", ns)
        resource = entry.find("d:resource", ns)
        patient_resource = resource.find("d:Patient", ns)
        if query_type == "id" or query_type == "all":
            id_resource = patient_resource.find("d:id", ns)
            patient["id"] = id_resource.attrib["value"]
        if query_type == "data" or query_type == "all":
            name = patient_resource.find("d:name", ns)
            if name != None:
                patient["name"] = ""
                family_name = name.find("d:family", ns)
                if family_name != None:
                    patient["name"] = family_name.attrib["value"]
                given_name = name.find("d:given", ns)
                if given_name != None:
                    patient["name"] = patient["name"] + " " + given_name.attrib["value"]
                patient["name"] = patient["name"].strip()
            telecom = patient_resource.find("d:telecom", ns)
            if telecom != None:
                value = telecom.find("d:value", ns)
                if value != None:
                    patient["telecom"] = value.attrib["value"]
            gender = patient_resource.find("d:gender", ns)
            if gender != None:
                if gender.attrib["value"] == "male":
                    patient["gender"] = "Nam"
                elif gender.attrib["value"] == "female":
                    patient["gender"] = "Nữ"
            birthdate = patient_resource.find("d:birthDate", ns)
            if birthdate != None:
                patient["birthdate"] = birthdate.attrib["value"]
            addresses = patient_resource.findall("d:address", ns)
            for address in addresses:
                addr_type = address.find("d:use", ns).attrib["value"]
                if addr_type == "home":
                    patient["home_address"] = (
                        address.find("d:line", ns).attrib["value"]
                        + ", "
                        + address.find("d:district", ns).attrib["value"]
                        + ", "
                        + address.find("d:city", ns).attrib["value"]
                    )
                elif addr_type == "work":
                    patient["work_address"] = (
                        address.find("d:line", ns).attrib["value"]
                        + ", "
                        + address.find("d:district", ns).attrib["value"]
                        + ", "
                        + address.find("d:city", ns).attrib["value"]
                    )
            patient["identifier"] = patient_identifier
            contact = patient_resource.find("d:contact", ns)
            if contact != None:
                relationship = contact.find("d:relationship", ns)
                if relationship != None:
                    patient["contact_relationship"] = relationship.find(
                        "d:text", ns
                    ).attrib["value"]
                name = contact.find("d:name", ns)
                if name != None:
                    patient["contact_name"] = (
                        name.find("d:family", ns).attrib["value"]
                        + " "
                        + name.find("d:given", ns).attrib["value"]
                    )
                telecom = contact.find("d:telecom", ns)
                if telecom != None:
                    patient["contact_telecom"] = telecom.find("d:value", ns).attrib[
                        "value"
                    ]
                address = contact.find("d:address", ns)
                if address != None:
                    patient["contact_address"] = (
                        address.find("d:line", ns).attrib["value"]
                        + ", "
                        + address.find("d:district", ns).attrib["value"]
                        + ", "
                        + address.find("d:city", ns).attrib["value"]
                    )
                gender = contact.find("d:gender", ns)
                if gender != None:
                    gender_value = gender.attrib["value"]
                    if gender_value == "male":
                        patient["contact_gender"] = "Nam"
                    else:
                        patient["contact_gender"] = "Nữ"
            # print(patient)
        return patient
    else:
        return None


def query_encounter(encounter_identifier, query_type):
    encounter = {}
    get_encounter = requests.get(
        fhir_server + "/Encounter?identifier=urn:trinhcongminh|" + encounter_identifier,
        headers={"Content-type": "application/xml"},
    )

    if get_encounter.status_code == 200 and "entry" in get_encounter.content.decode(
        "utf-8"
    ):
        get_root = ET.fromstring(get_encounter.content.decode("utf-8"))
        entry = get_root.find("d:entry", ns)
        resource = entry.find("d:resource", ns)
        encounter_resource = resource.find("d:Encounter", ns)
        if query_type == "meta" or query_type == "all":
            meta = encounter_resource.find("d:meta", ns)
            encounter["version"] = meta.find("d:versionId", ns).attrib["value"]
            encounter["last_updated"] = meta.find("d:lastUpdated", ns).attrib["value"]
        if query_type == "id" or query_type == "all":
            encounter["id"] = encounter_resource.find("d:id", ns).attrib["value"]
        if query_type == "data" or query_type == "all":
            encounter["encounter_identifier"] = encounter_identifier
            status = encounter_resource.find("d:status", ns)
            if status != None:
                encounter["encounter_status"] = status.attrib["value"]
            _class = encounter_resource.find("d:class", ns)
            if _class != None:
                encounter["encounter_class"] = _class.find("d:code", ns).attrib["value"]
            _type = encounter_resource.find("d:type", ns)
            if _type != None:
                encounter["encounter_type"] = _type.find("d:text", ns).attrib["value"]
            service_type = encounter_resource.find("d:serviceType", ns)
            if service_type != None:
                encounter["encounter_service"] = service_type.find("d:text", ns).attrib[
                    "value"
                ]
            priority = encounter_resource.find("d:priority", ns)
            if priority != None:
                encounter["encounter_priority"] = priority.find("d:text", ns).attrib[
                    "value"
                ]
            participant = encounter_resource.find("d:participant", ns)
            if participant != None:
                individual = participant.find("d:individual", ns)
                encounter["encounter_participant"] = individual.find(
                    "d:display", ns
                ).attrib["value"]
            period = encounter_resource.find("d:period", ns)
            if period != None:
                encounter["encounter_start"] = period.find("d:start", ns).attrib[
                    "value"
                ]
                end_date = None
                if period.find("d:end", ns) != None:
                    end_date = period.find("d:end", ns).attrib["value"]
                    encounter["encounter_end"] = getdatetime(end_date)
            length = encounter_resource.find("d:length", ns)
            if length != None:
                encounter["encounter_length"] = length.find("d:value", ns).attrib[
                    "value"
                ]
            reason_code = encounter_resource.find("d:reasonCode", ns)
            if reason_code != None:
                encounter["encounter_reason"] = reason_code.find("d:text", ns).attrib[
                    "value"
                ]

    return encounter


def query_service(service_identifier, query_type):
    service = {}
    get_service = requests.get(
        fhir_server
        + "/ServiceRequest?identifier=urn:trinhcongminh|"
        + service_identifier,
        headers={"Content-type": "application/xml"},
    )
    if get_service.status_code == 200 and "entry" in get_service.content.decode(
        "utf-8"
    ):
        get_root = ET.fromstring(get_service.content.decode("utf-8"))
        entry = get_root.find("d:entry", ns)
        resource = entry.find("d:resource", ns)
        service_resource = resource.find("d:ServiceRequest", ns)
        if query_type == "meta" or query_type == "all":
            meta = service_resource.find("d:meta", ns)
            service["version"] = meta.find("d:versionId", ns).attrib["value"]
            service["last_updated"] = meta.find("d:lastUpdated", ns).attrib["value"]
        if query_type == "id" or query_type == "all":
            service["id"] = service_resource.find("d:id", ns).attrib["value"]
        if query_type == "data" or query_type == "all":
            service["service_identifier"] = service_identifier
            status = service_resource.find("d:status", ns)
            if status != None:
                service["service_status"] = status.attrib["value"]
            intent = service_resource.find("d:intent", ns)
            if intent != None:
                service["service_intent"] = intent.attrib["value"]
            category = service_resource.find("d:category", ns)
            if category != None:
                service["service_category"] = category.find("d:text", ns).attrib[
                    "value"
                ]
            code = service_resource.find("d:code", ns)
            if code != None:
                service["service_code"] = code.find("d:text", ns).attrib["value"]
            occurrence = service_resource.find("d:occurrenceDateTime", ns)
            if occurrence != None:
                service["service_occurrence"] = get_date(occurrence.attrib["value"])
            authored_on = service_resource.find("d:authoredOn", ns)
            if authored_on != None:
                service["service_authored"] = get_date(authored_on.attrib["value"])
            requester = service_resource.find("d:requester", ns)
            if requester != None:
                display = requester.find("d:display", ns)
                if display != None:
                    service["service_requester"] = display.attrib["value"]
            performer = service_resource.find("d:performer", ns)
            if performer != None:
                display = performer.find("d:display", ns)
                if display != None:
                    service["service_performer"] = display.attrib["value"]
            reason_code = service_resource.find("d:reasonCode", ns)
            if reason_code != None:
                text = reason_code.find("d:text", ns)
                if text != None:
                    service["service_reason_code"] = text.attrib["value"]
            note = service_resource.find("d:note", ns)
            if note != None:
                service["service_note"] = note.find("d:text", ns).attrib["value"]
    return service


def query_condition(condition_identifier, query_type):
    condition = {}
    get_condition = requests.get(
        fhir_server + "/Condition?identifier=urn:trinhcongminh|" + condition_identifier,
        headers={"Content-type": "application/xml"},
    )

    if get_condition.status_code == 200 and "entry" in get_condition.content.decode(
        "utf-8"
    ):
        get_root = ET.fromstring(get_condition.content.decode("utf-8"))
        entry = get_root.find("d:entry", ns)
        resource = entry.find("d:resource", ns)
        condition_resource = resource.find("d:Condition", ns)
        if query_type == "meta" or query_type == "all":
            meta = condition_resource.find("d:meta", ns)
            condition["version"] = meta.find("d:versionId", ns).attrib["value"]
            condition["last_updated"] = meta.find("d:lastUpdated", ns).attrib["value"]
        if query_type == "id" or query_type == "all":
            condition["id"] = condition_resource.find("d:id", ns).attrib["value"]
        if query_type == "data" or query_type == "all":
            condition["condition_identifier"] = condition_identifier
            clinical_status = condition_resource.find("d:clinicalStatus", ns)
            if clinical_status != None:
                condition["condition_clinical_status"] = clinical_status.find(
                    "d:text", ns
                ).attrib["value"]
            verification_status = condition_resource.find("d:verificationStatus", ns)
            if verification_status != None:
                condition["condition_verification_status"] = verification_status.find(
                    "d:text", ns
                ).attrib["value"]
            category = condition_resource.find("d:category", ns)
            if category != None:
                condition["condition_category"] = category.find("d:text", ns).attrib[
                    "value"
                ]
            severity = condition_resource.find("d:severity", ns)
            if severity != None:
                condition["get_condition_severity_display"] = severity.find(
                    "d:text", ns
                ).attrib["value"]
            code = condition_resource.find("d:code", ns)
            if code != None:
                condition["condition_code"] = code.find("d:text", ns).attrib["value"]
            onset = condition_resource.find("d:onsetDateTime", ns)
            if onset != None:
                condition["condition_onset"] = onset.attrib["value"]
            abatement = condition_resource.find("d:abatementDateTime", ns)
            if abatement != None:
                condition["condition_abatement"] = abatement.attrib["value"]
            note = condition_resource.find("d:note", ns)
            if note != None:
                condition["condition_note"] = note.find("d:text", ns).attrib["value"]

    return condition


def query_observation(observation_identifier, query_type):
    observation = {}
    get_observation = requests.get(
        fhir_server
        + "/Observation?identifier=urn:trinhcongminh|"
        + observation_identifier,
        headers={"Content-type": "application/xml"},
    )
    if (
        get_observation.status_code == 200
        and "entry" in get_observation.content.decode("utf-8")
    ):
        get_root = ET.fromstring(get_observation.content.decode("utf-8"))
        entry = get_root.find("d:entry", ns)
        resource = entry.find("d:resource", ns)
        observation_resource = resource.find("d:Observation", ns)
        if query_type == "meta" or query_type == "all":
            meta = observation_resource.find("d:meta", ns)
            observation["version"] = meta.find("d:versionId", ns).attrib["value"]
            observation["last_updated"] = meta.find("d:lastUpdated", ns).attrib["value"]
        if query_type == "id" or query_type == "all":
            observation["id"] = observation_resource.find("d:id", ns).attrib["value"]
        if query_type == "data" or query_type == "all":
            observation["observation_identifier"] = observation_identifier
            status = observation_resource.find("d:status", ns)
            if status != None:
                observation["observation_status"] = status.attrib["value"]
            category = observation_resource.find("d:category", ns)
            if category != None:
                observation["observation_category"] = category.find(
                    "d:text", ns
                ).attrib["value"]
            code = observation_resource.find("d:code", ns)
            if code != None:
                observation["observation_code"] = code.find("d:text", ns).attrib[
                    "value"
                ]
            effective = observation_resource.find("d:effectiveDateTime", ns)
            if effective != None:
                observation["observation_effective"] = getdatetime(
                    effective.attrib["value"]
                )
            value_quantity = observation_resource.find("d:valueQuantity", ns)
            if value_quantity != None:
                observation["observation_value_quantity"] = value_quantity.find(
                    "d:value", ns
                ).attrib["value"]
                unit = value_quantity.find("d:unit", ns)
                if unit != None:
                    observation["observation_value_unit"] = unit.attrib["value"]
            reference_range = observation_resource.find("d:referenceRange", ns)
            if reference_range != None:
                observation["observation_reference_range"] = reference_range.find(
                    "d:text", ns
                ).attrib["value"]
    return observation


def query_procedure(procedure_identifier, query_type):
    procedure = {}
    get_procedure = requests.get(
        fhir_server + "/Procedure?identifier=urn:trinhcongminh|" + procedure_identifier,
        headers={"Content-type": "application/xml"},
    )
    if get_procedure.status_code == 200 and "entry" in get_procedure.content.decode(
        "utf-8"
    ):
        get_root = ET.fromstring(get_procedure.content.decode("utf-8"))
        entry = get_root.find("d:entry", ns)
        resource = entry.find("d:resource", ns)
        procedure_resource = resource.find("d:Procedure", ns)
        if query_type == "meta" or query_type == "all":
            meta = procedure_resource.find("d:meta", ns)
            procedure["version"] = meta.find("d:versionId", ns).attrib["value"]
            procedure["last_updated"] = meta.find("d:lastUpdated", ns).attrib["value"]
        if query_type == "id" or query_type == "all":
            procedure["id"] = procedure_resource.find("d:id", ns).attrib["value"]
        if query_type == "data" or query_type == "all":
            procedure["procedure_identifier"] = procedure_identifier
            status = procedure_resource.find("d:status", ns)
            if status != None:
                procedure["procedure_status"] = status.attrib["value"]
            category = procedure_resource.find("d:category", ns)
            if category != None:
                procedure["procedure_category"] = category.find("d:text", ns).attrib[
                    "value"
                ]
            code = procedure_resource.find("d:code", ns)
            if code != None:
                procedure["procedure_code"] = code.find("d:text", ns).attrib["value"]
            performed_datetime = procedure_resource.find("d:performedDateTime", ns)
            if performed_datetime != None:
                procedure["procedure_performed_datetime"] = getdatetime(
                    performed_datetime.attrib["value"]
                )
            performer = procedure_resource.find("d:performer", ns)
            if performer != None:
                actor = performer.find("d:actor", ns)
                display = actor.find("d:display", ns)
                if display != None:
                    procedure["procedure_performer"] = display.attrib["value"]
            reason_code = procedure_resource.find("d:reasonCode", ns)
            if reason_code != None:
                text = reason_code.find("d:text", ns)
                if text != None:
                    procedure["procedure_reason_code"] = text.attrib["value"]
            outcome = procedure_resource.find("d:outcome", ns)
            if outcome != None:
                text = outcome.find("d:text", ns)
                if text != None:
                    procedure["procedure_outcome"] = text.attrib["value"]
            complication = procedure_resource.find("d:complication", ns)
            if complication != None:
                text = complication.find("d:text", ns)
                if text != None:
                    procedure["procedure_complication"] = text.attrib["value"]
            follow_up = procedure_resource.find("d:followUp", ns)
            if follow_up != None:
                text = follow_up.find("d:text", ns)
                if text != None:
                    procedure["procedure_follow_up"] = text.attrib["value"]
            note = procedure_resource.find("d:note", ns)
            if note != None:
                text = note.find("d:text", ns)
                if text != None:
                    procedure["procedure_note"] = note.find("d:text", ns).attrib[
                        "value"
                    ]
    return procedure


def query_medication(medication_identifier, query_type):
    medication_statement = {}
    get_medication = requests.get(
        fhir_server
        + "/MedicationStatement?identifier=urn:trinhcongminh|"
        + medication_identifier,
        headers={"Content-type": "application/xml"},
    )
    if get_medication.status_code == 200 and "entry" in get_medication.content.decode(
        "utf-8"
    ):
        get_root = ET.fromstring(get_medication.content.decode("utf-8"))
        entry = get_root.find("d:entry", ns)
        resource = entry.find("d:resource", ns)
        medication_resource = resource.find("d:MedicationStatement", ns)
        if query_type == "meta" or query_type == "all":
            meta = medication_resource.find("d:meta", ns)
            medication_statement["version"] = meta.find("d:versionId", ns).attrib[
                "value"
            ]
            medication_statement["last_updated"] = meta.find(
                "d:lastUpdated", ns
            ).attrib["value"]
        if query_type == "id" or query_type == "all":
            medication_statement["id"] = medication_resource.find("d:id", ns).attrib[
                "value"
            ]
        if query_type == "data" or query_type == "all":
            medication_statement["medication_identifier"] = medication_identifier
            medication = medication_resource.find("d:medicationCodeableConcept", ns)
            if medication != None:
                medication_statement["medication_medication"] = medication.find(
                    "d:text", ns
                ).attrib["value"]
            effective = medication_resource.find("d:effectiveDateTime", ns)
            if effective != None:
                medication_statement["medication_effective"] = get_date(
                    effective.attrib["value"]
                )
            date_asserted = medication_resource.find("d:dateAsserted", ns)
            if date_asserted != None:
                medication_statement["medication_date_asserted"] = get_date(
                    date_asserted.attrib["value"]
                )
            reason_code = medication_resource.find("d:reasonCode", ns)
            if reason_code != None:
                medication_statement["medication_reason_code"] = reason_code.find(
                    "d:text", ns
                ).attrib["value"]
            dosage = medication_resource.find("d:dosage", ns)
            if dosage != None:
                additional_instruction = dosage.find("d:additionalInstruction", ns)
                if additional_instruction != None:
                    medication_statement[
                        "dosage_additional_instruction"
                    ] = additional_instruction.find("d:text", ns).attrib["value"]
                patient_instruction = dosage.find("d:patientInstruction", ns)
                if patient_instruction != None:
                    medication_statement[
                        "dosage_patient_instruction"
                    ] = patient_instruction.find("d:text", ns).attrib["value"]
                timing = dosage.find("d:timing", ns)
                if timing != None:
                    repeat = timing.find("d:repeat", ns)
                    if repeat != None:
                        duration = repeat.find("d:duration", ns)
                        if duration != None:
                            duration_value = duration.attrib["value"]
                            duration_max = repeat.find("d:durationMax", ns)
                            if duration_max != None:
                                duration_value = (
                                    duration.attrib["value"]
                                    + "-"
                                    + duration_max.attrib["value"]
                                )
                            duration_unit = repeat.find("d:durationUnit", ns).attrib[
                                "value"
                            ]
                            medication_statement["dosage_duration"] = duration_value
                            medication_statement["dosage_duration_unit"] = duration_unit
                        frequency = repeat.find("d:frequency", ns)
                        if frequency != None:
                            medication_statement["dosage_frequency"] = (
                                frequency.attrib["value"] + " lần"
                            )
                        period = repeat.find("d:period", ns)
                        if period != None:
                            period_value = period.attrib["value"]
                            period_max = repeat.find("d:periodMax", ns)
                            if period_max != None:
                                period_value = (
                                    period.attrib["value"]
                                    + "-"
                                    + period_max.attrib["value"]
                                )
                            period_unit = repeat.find("d:periodUnit", ns)
                            medication_statement["dosage_period"] = period_value
                            medication_statement["dosage_period_value"] = period_unit
                        when = repeat.find("d:when", ns)
                        if when != None:
                            medication_statement["dosage_when"] = when.attrib["value"]
                        offset = repeat.find("d:offset", ns)
                        if offset != None:
                            medication_statement["dosage_offset"] = offset.attrib[
                                "value"
                            ]
                route = dosage.find("d:route", ns)
                if route != None:
                    medication_statement["dosage_route"] = route.find(
                        "d:text", ns
                    ).attrib["value"]
                dose_and_rate = dosage.find("d:doseAndRate", ns)
                if dose_and_rate != None:
                    dose_quantity = dose_and_rate.find("d:doseQuantity", ns)
                    quantity = dose_quantity.find("d:value", ns).attrib["value"]
                    unit = dose_quantity.find("d:unit", ns).attrib["value"]
                    medication_statement["dosage_quantity"] = quantity + " " + unit
    return medication_statement


def query_practitioner(practitioner_identifier, query_type):
    practitioner = {}
    get_practitioner = requests.get(
        fhir_server
        + "/Practitioner?identifier=urn:trinhcongminh|"
        + practitioner_identifier,
        headers={"Content-type": "application/xml"},
    )
    # print(get_practitioner.content.decode('utf-8'))
    if (
        get_practitioner.status_code == 200
        and "entry" in get_practitioner.content.decode("utf-8")
    ):
        get_root = ET.fromstring(get_practitioner.content.decode("utf-8"))
        entry = get_root.find("d:entry", ns)
        resource = entry.find("d:resource", ns)
        practitioner_resource = resource.find("d:Practitioner", ns)
        if query_type == "meta" or query_type == "all":
            meta = practitioner_resource.find("d:meta", ns)
            practitioner["version"] = meta.find("d:versionId", ns).attrib["value"]
            practitioner["last_updated"] = meta.find("d:lastUpdated", ns).attrib[
                "value"
            ]
        if query_type == "id" or query_type == "all":
            print("have id")
            practitioner["id"] = practitioner_resource.find("d:id", ns).attrib["value"]
        if query_type == "data" or query_type == "all":
            practitioner["identifier"] = practitioner_identifier
            practitioner_name = practitioner_resource.find("d:name", ns)
            if practitioner_name != None:
                practitioner["name"] = (
                    practitioner_name.find("d:family", ns).attrib["value"]
                    + " "
                    + practitioner_name.find("d:given", ns).attrib["value"]
                )
            practitioner_telecom = practitioner_resource.find("d:telecom", ns)
            if practitioner_telecom != None:
                practitioner["telecom"] = practitioner_telecom.find(
                    "d:value", ns
                ).attrib["value"]
            practitioner_gender = practitioner_resource.find("d:gender", ns)
            if practitioner_gender != None:
                if practitioner_gender.attrib["value"] == "male":
                    practitioner["gender"] = "Nam"
                else:
                    practitioner["gender"] = "Nữ"
            practitioner_birthdate = practitioner_resource.find("d:birthDate", ns)
            if practitioner_birthdate != None:
                practitioner["birthdate"] = practitioner_birthdate.attrib["value"]
            practitioner_qualification = practitioner_resource.find(
                "d:qualification", ns
            )
            if practitioner_qualification != None:
                practitioner_qualification_code = practitioner_qualification.find(
                    "d:code", ns
                )
                practitioner["qualification"] = practitioner_qualification_code.find(
                    "d:text", ns
                ).attrib["value"]
    return practitioner


def query_diagnostic_report(diagnostic_report_identifier, query_type):
    diagnostic_report = {}
    get_diagnostic_report = requests.get(
        fhir_server
        + "/DiagnosticReport?identifier=urn:trinhcongminh|"
        + diagnostic_report_identifier,
        headers={"Content-type": "application/xml"},
    )
    if (
        get_diagnostic_report.status_code == 200
        and "entry" in get_diagnostic_report.content.decode("utf-8")
    ):
        get_root = ET.fromstring(get_diagnostic_report.content.decode("utf-8"))
        entry = get_root.find("d:entry", ns)
        resource = entry.find("d:resource", ns)
        diagnostic_report_resource = resource.find("d:DiagnosticReport", ns)
        if query_type == "meta" or query_type == "all":
            meta = diagnostic_report_resource.find("d:meta", ns)
            diagnostic_report["version"] = meta.find("d:versionId", ns).attrib["value"]
            diagnostic_report["last_updated"] = getdatetime(
                meta.find("d:lastUpdated", ns).attrib["value"]
            )
        if query_type == "id" or query_type == "all":
            id_ = diagnostic_report_resource.find("d:id", ns)
            diagnostic_report["id"] = id_.attrib["value"]
        if query_type == "data" or query_type == "all":
            diagnostic_report["diagnostic_identifier"] = diagnostic_report_identifier
            status = diagnostic_report_resource.find("d:status", ns)
            if status != None:
                diagnostic_report["diagnostic_status"] = status.attrib["value"]
            category = diagnostic_report_resource.find("d:category", ns)
            if category != None:
                diagnostic_report["diagnostic_category"] = category.find(
                    "d:text", ns
                ).attrib["value"]
            code = diagnostic_report_resource.find("d:code", ns)
            if code != None:
                diagnostic_report["diagnostic_code"] = code.find("d:text", ns).attrib[
                    "value"
                ]
            effective = diagnostic_report_resource.find("d:effectiveDateTime", ns)
            if effective != None:
                diagnostic_report["diagnostic_effective"] = effective.attrib["value"]
            performer = diagnostic_report_resource.find("d:performer", ns)
            if performer != None:
                display = performer.find("d:display", ns)
                if display != None:
                    diagnostic_report["performer"] = display.attrib["value"]
            conclusion = diagnostic_report_resource.find("d:conclusion", ns)
            if conclusion != None:
                diagnostic_report["diagnostic_conclusion"] = conclusion.attrib["value"]
    return diagnostic_report


def query_allergy(allergy_identifier, query_type):
    allergy = {}
    get_allergy = requests.get(
        fhir_server
        + "/AllergyIntolerance?identifier=urn:trinhcongminh|"
        + allergy_identifier,
        headers={"Content-type": "application/xml"},
    )
    if get_allergy.status_code == 200 and "entry" in get_allergy.content.decode(
        "utf-8"
    ):
        get_root = ET.fromstring(get_allergy.content.decode("utf-8"))
        entry = get_root.find("d:entry", ns)
        resource = entry.find("d:resource", ns)
        allergy_resource = resource.find("d:AllergyIntolerance", ns)
        if query_type == "all" or query_type == "meta":
            meta = allergy_resource.find("d:meta", ns)
            allergy["version"] = meta.find("d:versionId", ns).attrib["value"]
            allergy["last_updated"] = getdatetime(
                meta.find("d:lastUpdated").attrib["value"]
            )
        if query_type == "all" or query_type == "id":
            id_ = allergy_resource.find("d:id", ns)
            allergy["id"] = id_.attrib["value"]
        if query_type == "all" or query_type == "data":
            allergy["allergy_identifier"] = allergy_identifier
            clinical_status = allergy_resource.find("d:clinicalStatus", ns)
            if clinical_status != None:
                allergy["allergy_clinical_status"] = clinical_status.find(
                    "d:text", ns
                ).attrib["value"]
            verification_status = allergy_resource.find("d:verificationStatus", ns)
            if verification_status != None:
                allergy["allergy_verification_status"] = verification_status.find(
                    "d:text", ns
                ).attrib["value"]
            category = allergy_resource.find("d:category", ns)
            if category != None:
                allergy["allergy_category"] = category.find("d:text", ns).attrib[
                    "value"
                ]
            criticality = allergy_resource.find("d:criticality", ns)
            if criticality != None:
                allergy["get_allergy_criticality_display"] = criticality.find(
                    "d:text", ns
                ).attrib["value"]
            code = allergy_resource.find("d:code", ns)
            if code != None:
                allergy["allergy_code"] = code.find("d:text", ns).attrib["value"]
            onset = allergy_resource.find("d:onsetDateTime", ns)
            if onset != None:
                allergy["allergy_onset"] = onset.attrib["value"]
            last_occurrence = allergy_resource.find("d:lastOccurrence")
            if last_occurrence != None:
                allergy["allergy_last_occurrence"] = last_occurrence.attrib["value"]
            reaction = allergy_resource.find("d:reaction", ns)
            if reaction != None:
                substance = reaction.find("d:substance", ns)
                if substance != None:
                    allergy["allergy_reaction_substance"] = substance.find(
                        "d:text", ns
                    ).attrib["value"]
                manifestation = reaction.find("d:manifestation", ns)
                if manifestation != None:
                    allergy["allergy_reaction_manifestation"] = manifestation.find(
                        "d:text", ns
                    ).attrib["value"]
                severity = allergy_resource.find("d:severity", ns)
                if severity != None:
                    allergy["get_allergy_reaction_severity_display"] = severity.find(
                        "d:severity", ns
                    ).attrib["value"]
    return allergy


def query_encounter_history(encounter_id, version, query_type):
    encounter = {}
    get_encounter = requests.get(
        fhir_server + "/Encounter/" + encounter_id + "/_history/" + version,
        headers={"Content-type": "application/xml"},
    )
    if get_encounter.status_code == 200:
        # get_root = ET.fromstring(get_encounter.content.decode('utf-8'))
        # entry = get_root.find('d:entry', ns)
        # resource = entry.find('d:resource', ns)
        encounter_resource = ET.fromstring(get_encounter.content.decode("utf-8"))
        if query_type == "meta" or query_type == "all":
            meta = encounter_resource.find("d:meta", ns)
            encounter["version"] = meta.find("d:versionId", ns).attrib["value"]
            encounter["last_updated"] = getdatetime(
                meta.find("d:lastUpdated", ns).attrib["value"]
            )
        if query_type == "id" or query_type == "all":
            encounter["id"] = encounter_resource.find("d:id", ns).attrib["value"]
        if query_type == "data" or query_type == "all":
            # encounter['encounter_identifier'] = encounter_identifier
            status = encounter_resource.find("d:status", ns)
            if status != None:
                encounter["encounter_status"] = status.attrib["value"]
            _class = encounter_resource.find("d:class", ns)
            if _class != None:
                encounter["encounter_class"] = _class.find("d:code", ns).attrib["value"]
            _type = encounter_resource.find("d:type", ns)
            if _type != None:
                encounter["encounter_type"] = _type.find("d:text", ns).attrib["value"]
            service_type = encounter_resource.find("d:serviceType", ns)
            if service_type != None:
                encounter["encounter_service"] = service_type.find("d:text", ns).attrib[
                    "value"
                ]
            # priority = encounter_resource.find('d:priority', ns)
            # if priority:
            #     encounter['encounter_priority'] = priority.find('d:text', ns).attrib['value']
            period = encounter_resource.find("d:period", ns)
            if period != None:
                start_date = period.find("d:start", ns).attrib["value"]
                encounter["encounter_start"] = getdatetime(start_date)
                end_date = None
                if period.find("d:end", ns) != None:
                    end_date = period.find("d:end", ns).attrib["value"]
                    encounter["encounter_end"] = getdatetime(end_date)
            length = encounter_resource.find("d:length", ns)
            if length != None:
                encounter["encounter_length"] = length.find("d:value", ns).attrib[
                    "value"
                ]
            reason_code = encounter_resource.find("d:reasonCode", ns)
            if reason_code != None:
                encounter["encounter_reason"] = reason_code.find("d:text", ns).attrib[
                    "value"
                ]
    return encounter


def query_service_history(service_id, version, query_type):
    service = {}
    get_service = requests.get(
        fhir_server + "/ServiceRequest/" + service_id + "/_history" + version,
        headers={"Content-type": "application/xml"},
    )
    if get_service.status_code == 200:
        # get_root = ET.fromstring(get_service.content.decode('utf-8'))
        # entry = get_root.find('d:entry', ns)
        # resource = entry.find('d:resource', ns)
        service_resource = ET.fromstring(get_service.content.decode("utf-8"))
        if query_type == "meta" or query_type == "all":
            meta = service_resource.find("d:meta", ns)
            service["version"] = meta.find("d:versionId", ns).attrib["value"]
            service["last_updated"] = getdatetime(
                meta.find("d:lastUpdated", ns).attrib["value"]
            )
        if query_type == "id" or query_type == "all":
            service["id"] = service_resource.find("d:id", ns).attrib["value"]
        if query_type == "data" or query_type == "all":
            # service['service_identifier'] = service_identifier
            status = service_resource.find("d:status", ns)
            if status != None:
                service['get_service_status_display'] = status.attrib['value']
            intent = service_resource.find('d:intent', ns)
            if intent != None:
                service["service_intent"] = intent.attrib["value"]
            category = service_resource.find("d:category", ns)
            if category != None:
                service["service_category"] = category.find("d:text", ns).attrib[
                    "value"
                ]
            code = service_resource.find("d:code", ns)
            if code != None:
                service["service_code"] = code.find("d:text", ns).attrib["value"]
            occurrence = service_resource.find("d:occurrenceDateTime", ns)
            if occurrence != None:
                service["service_occurrence"] = get_date(occurrence.attrib["value"])
            authored_on = service_resource.find("d:authoredOn", ns)
            if authored_on != None:
                service["service_authored"] = get_date(authored_on.attrib["value"])
            note = service_resource.find("d:note", ns)
            if note != None:
                service["service_note"] = note.find("d:text", ns).attrib["value"]
    return service


def query_condition_history(condition_id, version, query_type):
    condition = {}
    get_condition = requests.get(
        fhir_server + "/Condition/" + condition_id + "/_history/" + version,
        headers={"Content-type": "application/xml"},
    )
    if get_condition.status_code == 200:
        # get_root = ET.fromstring(get_condition.content.decode('utf-8'))
        # entry = get_root.find('d:entry', ns)
        # resource = entry.find('d:resource', ns)
        condition_resource = ET.fromstring(get_condition.content.decode("utf-8"))
        if query_type == "meta" or query_type == "all":
            meta = condition_resource.find("d:meta", ns)
            condition["version"] = meta.find("d:versionId", ns).attrib["value"]
            condition["last_updated"] = getdatetime(
                meta.find("d:lastUpdated", ns).attrib["value"]
            )
        if query_type == "id" or query_type == "all":
            condition["id"] = condition_resource.find("d:id", ns).attrib["value"]
        if query_type == "data" or query_type == "all":
            # condition['condition_identifier'] = condition_identifier
            clinical_status = condition_resource.find("d:clinicalStatus", ns)
            if clinical_status != None:
                condition["condition_clinical_status"] = clinical_status.find(
                    "d:text", ns
                ).attrib["value"]
            verification_status = condition_resource.find("d:verificationStatus", ns)
            if verification_status != None:
                condition["condition_verification_status"] = verification_status.find(
                    "d:text", ns
                ).attrib["value"]
            category = condition_resource.find("d:category", ns)
            if category != None:
                condition["condition_category"] = category.find("d:text", ns).attrib[
                    "value"
                ]
            severity = condition_resource.find("d:severity", ns)
            if severity != None:
                condition["condition_severity"] = severity.find("d:text", ns).attrib[
                    "value"
                ]
            code = condition_resource.find("d:code", ns)
            if code != None:
                condition["condition_code"] = code.find("d:text", ns).attrib["value"]
            onset = condition_resource.find("d:onsetDateTime", ns)
            if onset != None:
                condition["condition_onset"] = onset.attrib["value"]
            abatement = condition_resource.find("d:abatementDateTime", ns)
            if abatement != None:
                condition["condition_abatement"] = abatement.attrib["value"]
            note = condition_resource.find("d:note", ns)
            if note != None:
                condition["condition_note"] = note.find("d:text", ns).attrib["value"]

    return condition


def query_observation_history(observation_id, version, query_type):
    observation = {}
    get_observation = requests.get(
        fhir_server + "/Observation/" + observation_id + "/_history/" + version,
        headers={"Content-type": "application/xml"},
    )
    if get_observation.status_code == 200:
        # get_root = ET.fromstring(get_observation.content.decode('utf-8'))
        # entry = get_root.find('d:entry', ns)
        # resource = entry.find('d:resource', ns)
        observation_resource = ET.fromstring(get_observation.content.decode("utf-8"))
        if query_type == "meta" or query_type == "all":
            meta = observation_resource.find("d:meta", ns)
            observation["version"] = meta.find("d:versionId", ns).attrib["value"]
            observation["last_updated"] = getdatetime(
                meta.find("d:lastUpdated", ns).attrib["value"]
            )
        if query_type == "id" or query_type == "all":
            observation["id"] = observation_resource.find("d:id", ns).attrib["value"]
        if query_type == "data" or query_type == "all":
            # observation['observation_identifier'] = observation_identifier
            status = observation_resource.find("d:status", ns)
            if status != None:
                observation["observation_status"] = status.attrib["value"]
            category = observation_resource.find("d:category", ns)
            if category != None:
                observation["observation_category"] = category.find(
                    "d:text", ns
                ).attrib["value"]
            code = observation_resource.find("d:code", ns)
            if code != None:
                observation["observation_code"] = code.find("d:text", ns).attrib[
                    "value"
                ]
            effective = observation_resource.find("d:effectiveDateTime", ns)
            if effective != None:
                observation["observation_effective"] = getdatetime(
                    effective.attrib["value"]
                )
            value_quantity = observation_resource.find("d:valueQuantity", ns)
            if value_quantity != None:
                observation["observation_value_quantity"] = value_quantity.find(
                    "d:value", ns
                ).attrib["value"]
                unit = value_quantity.find("d:unit", ns)
                if unit != None:
                    observation["observation_value_unit"] = unit.attrib["value"]
            reference_range = observation_resource.find("d:referenceRange", ns)
            if reference_range != None:
                observation["observation_reference_range"] = reference_range.find(
                    "d:text", ns
                ).attrib["value"]

    return observation


def query_procedure_history(procedure_id, version, query_type):
    procedure = {}
    get_procedure = requests.get(
        fhir_server + "/Procedure/" + procedure_id + "/_history/" + version,
        headers={"Content-type": "application/xml"},
    )
    if get_procedure.status_code == 200:
        # get_root = ET.fromstring(get_procedure.content.decode('utf-8'))
        # entry = get_root.find('d:entry', ns)
        # resource = entry.find('d:resource', ns)
        procedure_resource = ET.fromstring(get_procedure.content.decode("utf-8"))
        if query_type == "meta" or query_type == "all":
            meta = procedure_resource.find("d:meta", ns)
            procedure["version"] = meta.find("d:versionId", ns).attrib["value"]
            procedure["last_updated"] = getdatetime(
                meta.find("d:lastUpdated", ns).attrib["value"]
            )
        if query_type == "id" or query_type == "all":
            procedure["id"] = procedure_resource.find("d:id", ns).attrib["value"]
        if query_type == "data" or query_type == "all":
            # procedure['procedure_identifier'] = procedure_identifier
            status = procedure_resource.find("d:status", ns)
            if status != None:
                procedure["procedure_status"] = status.attrib["value"]
            category = procedure_resource.find("d:category", ns)
            if category != None:
                procedure["procedure_category"] = category.find("d:text", ns).attrib[
                    "value"
                ]
            code = procedure_resource.find("d:code", ns)
            if code != None:
                procedure["procedure_code"] = code.find("d:text", ns).attrib["value"]
            performed_datetime = procedure_resource.find("d:performedDateTime", ns)
            if performed_datetime != None:
                procedure["procedure_performed_datetime"] = getdatetime(
                    performed_datetime.attrib["value"]
                )
            reason_code = procedure_resource.find("d:reasonCode", ns)
            if reason_code != None:
                procedure["procedure_reason_code"] = reason_code.find(
                    "d:text", ns
                ).attrib["value"]
            outcome = procedure_resource.find("d:outcome", ns)
            if outcome != None:
                procedure["procedure_outcome"] = outcome.find("d:text", ns).attrib[
                    "value"
                ]
            complication = procedure_resource.find("d:complication", ns)
            if complication != None:
                procedure["procedure_complication"] = complication.find(
                    "d:text", ns
                ).attrib["value"]
            follow_up = procedure_resource.find("d:followUp", ns)
            if follow_up != None:
                procedure["procedure_follow_up"] = follow_up.find("d:text", ns).attrib[
                    "value"
                ]
            note = procedure_resource.find("d:note", ns)
            if note != None:
                procedure["procedure_note"] = note.find("d:text", ns).attrib["value"]
    return procedure


def query_medication_history(medication_id, version, query_type):
    medication_statement = {}
    get_medication = requests.get(
        fhir_server + "/MedicationStatement/" + medication_id + "/_history/" + version,
        headers={"Content-type": "application/xml"},
    )
    if get_medication.status_code == 200:
        # get_root = ET.fromstring(get_medication.content.decode('utf-8'))
        # entry = get_root.find('d:entry', ns)
        # resource = entry.find('d:resource', ns)
        medication_resource = ET.fromstring(get_medication.content.decode("utf-8"))
        if query_type == "meta" or query_type == "all":
            meta = medication_resource.find("d:meta", ns)
            medication_statement["version"] = meta.find("d:versionId", ns).attrib[
                "value"
            ]
            medication_statement["last_updated"] = getdatetime(
                meta.find("d:lastUpdated", ns).attrib["value"]
            )
        if query_type == "id" or query_type == "all":
            medication_statement["id"] = medication_resource.find("d:id", ns).attrib[
                "value"
            ]
        if query_type == "data" or query_type == "all":
            # medication_statement['medication_identifier'] = medication_identifier
            medication = medication_resource.find("d:medicationCodeableConcept", ns)
            if medication:
                medication_statement["medication_medication"] = medication.find(
                    "d:text", ns
                ).attrib["value"]
            effective = medication_resource.find("d:effectiveDateTime", ns)
            if effective != None:
                medication_statement["medication_effective"] = get_date(
                    effective.attrib["value"]
                )
            date_asserted = medication_resource.find("d:dateAsserted", ns)
            if date_asserted != None:
                medication_statement["medication_date_asserted"] = get_date(
                    date_asserted.attrib["value"]
                )
            reason_code = medication_resource.find("d:reasonCode", ns)
            if reason_code != None:
                medication_statement["medication_reason_code"] = reason_code.find(
                    "d:text", ns
                ).attrib["value"]
            dosage = medication_resource.find("d:dosage", ns)
            if dosage:
                additional_instruction = dosage.find("d:additionalInstruction", ns)
                if additional_instruction:
                    medication_statement[
                        "dosage_additional_instruction"
                    ] = additional_instruction.find("d:text", ns).attrib["value"]
                patient_instruction = dosage.find("d:patientInstruction", ns)
                if patient_instruction:
                    medication_statement[
                        "dosage_patient_instruction"
                    ] = patient_instruction.find("d:text", ns).attrib["value"]
                timing = dosage.find("d:timing", ns)
                if timing:
                    repeat = timing.find("d:repeat", ns)
                    if repeat:
                        duration = repeat.find("d:duration", ns)
                        if duration != None:
                            duration_value = duration.attrib["value"]
                            duration_max = repeat.find("d:durationMax", ns)
                            if duration_max != None:
                                duration_value = (
                                    duration.attrib["value"]
                                    + "-"
                                    + duration_max.attrib["value"]
                                )
                            duration_unit = repeat.find("d:durationUnit", ns).attrib[
                                "value"
                            ]
                            medication_statement["dosage_duration"] = duration_value
                            medication_statement["dosage_duration_unit"] = duration_unit
                        frequency = repeat.find("d:frequency", ns)
                        if frequency != None:
                            medication_statement["dosage_frequency"] = (
                                frequency.attrib["value"] + " lần"
                            )
                        period = repeat.find("d:period", ns)
                        if period != None:
                            period_value = period.attrib["value"]
                            period_max = repeat.find("d:periodMax", ns)
                            if period_max != None:
                                period_value = (
                                    period.attrib["value"]
                                    + "-"
                                    + period_max.attrib["value"]
                                )
                            period_unit = repeat.find("d:periodUnit", ns)
                            medication_statement["dosage_period"] = period_value
                            medication_statement["dosage_period_value"] = period_unit
                        when = repeat.find("d:when", ns)
                        if when != None:
                            medication_statement["dosage_when"] = when.attrib["value"]
                        offset = repeat.find("d:offset", ns)
                        if offset != None:
                            medication_statement["dosage_offset"] = offset.attrib[
                                "value"
                            ]
                route = dosage.find("d:route", ns)
                if route:
                    medication_statement["dosage_route"] = route.find(
                        "d:text", ns
                    ).attrib["value"]
                dose_and_rate = dosage.find("d:doseAndRate", ns)
                if dose_and_rate:
                    dose_quantity = dose_and_rate.find("d:doseQuantity", ns)
                    quantity = dose_quantity.find("d:value", ns).attrib["value"]
                    unit = dose_quantity.find("d:unit", ns).attrib["value"]
                    medication_statement["dosage_quantity"] = quantity + " " + unit
    return medication_statement


def query_practitioner_history(practitioner_identifier, query_type):
    practitioner = {}
    get_practitioner = requests.get(
        fhir_server
        + "/Practitioner?identifier=urn:trinhcongminh|"
        + practitioner_identifier,
        headers={"Content-type": "application/xml"},
    )
    if get_practitioner.status_code == 200:
        # get_root = ET.fromstring(get_practitioner.content.decode('utf-8'))
        # entry = get_root.find('d:entry', ns)
        # resource = entry.find('d:resource', ns)
        practitioner_resource = ET.fromstring(get_practitioner.content.decode("utf-8"))
        if query_type == "meta" or query_type == "all":
            meta = practitioner_resource.find("d:meta", ns)
            practitioner["version"] = meta.find("d:versionId", ns).attrib["value"]
            practitioner["last_updated"] = getdatetime(
                meta.find("d:lastUpdated", ns).attrib["value"]
            )
        if query_type == "id" or query_type == "all":
            practitioner["id"] = practitioner_resource.find("d:id", ns).attrib["value"]
        if query_type == "data" or query_type == "all":
            practitioner["identifier"] = practitioner_identifier
            practitioner_name = practitioner_resource.find("d:name", ns)
            if practitioner_name:
                practitioner["name"] = (
                    practitioner_name.find("d:family", ns).attrib["value"]
                    + " "
                    + practitioner_name.find("d:given", ns).attrib["value"]
                )
            practitioner_telecom = practitioner_resource.find("d:telecom", ns)
            if practitioner_telecom:
                practitioner["telecom"] = practitioner_telecom.find(
                    "d:value", ns
                ).attrib["value"]
            practitioner_gender = practitioner_resource.find("d:gender", ns)
            if practitioner_gender != None:
                if practitioner_gender.attrib["value"] == "male":
                    practitioner["gender"] = "Nam"
                else:
                    practitioner["gender"] = "Nữ"
            practitioner_birthdate = practitioner_resource.find("d:birthDate", ns)
            if practitioner_birthdate != None:
                practitioner["birthdate"] = practitioner_birthdate.attrib["value"]
            practitioner_qualification = practitioner_resource.find(
                "d:qualification", ns
            )
            if practitioner_qualification:
                practitioner_qualification_code = practitioner_qualification.find(
                    "d:code", ns
                )
                practitioner["qualification"] = practitioner_qualification_code.find(
                    "d:text", ns
                ).attrib["value"]
    return practitioner


def query_diagnostic_report_history(diagnostic_report_id, version, query_type):
    diagnostic_report = {}
    get_diagnostic_report = requests.get(
        fhir_server
        + "/DiagnosticReport/"
        + diagnostic_report_id
        + "/_history/"
        + version,
        headers={"Content-type": "application/xml"},
    )
    if get_diagnostic_report.status_code == 200:
        # get_root = ET.fromstring(get_diagnostic_report.content.decode('utf-8'))
        # entry = get_root.find('d:entry', ns)
        # resource = entry.find('d:resource', ns)
        diagnostic_report_resource = ET.fromstring(
            get_diagnostic_report.content.decode("utf-8")
        )
        if query_type == "meta" or query_type == "all":
            meta = diagnostic_report_resource.find("d:meta", ns)
            diagnostic_report["version"] = meta.find("d:versionId", ns).attrib["value"]
            diagnostic_report["last_updated"] = getdatetime(
                meta.find("d:lastUpdated", ns).attrib["value"]
            )
        if query_type == "id" or query_type == "all":
            id_ = diagnostic_report_resource.find("d:id", ns)
            diagnostic_report["id"] = id_.attrib["value"]
        if query_type == "data" or query_type == "all":
            # diagnostic_report['diagnostic_identifier'] = diagnostic_report_identifier
            status = diagnostic_report_resource.find("d:status", ns)
            if status != None:
                diagnostic_report["diagnostic_status"] = status.attrib["value"]
            category = diagnostic_report_resource.find("d:category", ns)
            if category:
                diagnostic_report["diagnostic_category"] = category.find(
                    "d:text", ns
                ).attrib["value"]
            code = diagnostic_report_resource.find("d:code", ns)
            if code:
                diagnostic_report["diagnostic_code"] = code.find("d:text", ns).attrib[
                    "value"
                ]
            effective = diagnostic_report_resource.find("d:effectiveDateTime", ns)
            if effective != None:
                diagnostic_report["diagnostic_effective"] = effective.attrib["value"]
            conclusion = diagnostic_report_resource.find("d:conclusion", ns)
            if conclusion != None:
                diagnostic_report["diagnostic_conclusion"] = conclusion.attrib["value"]
    return diagnostic_report


def query_allergy_history(allergy_id, version, query_type):
    allergy = {}
    get_allergy = requests.get(
        fhir_server + "/AllergyIntolerance/" + allergy_id + "/_history/" + version,
        headers={"Content-type": "application/xml"},
    )
    if get_allergy.status_code == 200:
        allergy_resource = ET.fromstring(get_allergy.content.decode("utf-8"))
        if query_type == "all" or query_type == "meta":
            meta = allergy_resource.find("d:meta", ns)
            allergy["version"] = meta.find("d:versionId", ns).attrib["value"]
            allergy["last_updated"] = getdatetime(
                meta.find("d:lastUpdated", ns).attrib["value"]
            )
        if query_type == "all" or query_type == "id":
            id_ = allergy_resource.find("d:id", ns)
            allergy["id"] = id_.attrib["value"]
        if query_type == "all" or query_type == "data":
            clinical_status = allergy_resource.find("d:clinicalStatus", ns)
            if clinical_status:
                allergy["allergy_clinical_status"] = clinical_status.find(
                    "d:text", ns
                ).attrib["value"]
            verification_status = allergy_resource.find("d:verificationStatus", ns)
            if verification_status:
                allergy["allergy_verification_status"] = verification_status.find(
                    "d:text", ns
                ).attrib["value"]
            category = allergy_resource.find("d:category", ns)
            if category:
                allergy["allergy_category"] = category.find("d:text", ns).attrib[
                    "value"
                ]
            criticality = allergy_resource.find("d:criticality", ns)
            if criticality:
                allergy["allergy_criticality"] = criticality.find("d:text", ns).attrib[
                    "value"
                ]
            code = allergy_resource.find("d:code", ns)
            if code:
                allergy["allergy_code"] = code.find("d:text", ns).attrib["value"]
            onset = allergy_resource.find("d:onsetDateTime", ns)
            if onset != None:
                allergy["allergy_onset"] = onset.attrib["value"]
            last_occurrence = allergy_resource.find("d:lastOccurrence")
            if last_occurrence != None:
                allergy["allergy_last_occurrence"] = last_occurrence.attrib["value"]
            reaction = allergy_resource.find("d:reaction", ns)
            if reaction:
                substance = reaction.find("d:substance", ns)
                if substance:
                    allergy["allergy_reaction_substance"] = substance.find(
                        "d:text", ns
                    ).attrib["value"]
                manifestation = reaction.find("d:manifestation", ns)
                if manifestation:
                    allergy["allergy_reaction_manifestation"] = manifestation.find(
                        "d:text", ns
                    ).attrib["value"]
                severity = allergy_resource.find("d:severity", ns)
                if severity:
                    allergy["allergy_reaction_severity"] = severity.find(
                        "d:severity", ns
                    ).attrib["value"]
    return allergy


def get_encounter(encounter_resource, query_type):
    encounter = {}
    if query_type == "meta" or query_type == "all":
        meta = encounter_resource.find("d:meta", ns)
        encounter["version"] = meta.find("d:versionId", ns).attrib["value"]
        encounter["last_updated"] = meta.find("d:lastUpdated", ns).attrib["value"]
    if query_type == "id" or query_type == "all":
        encounter["id"] = encounter_resource.find("d:id", ns).attrib["value"]
    if query_type == "data" or query_type == "all":
        # encounter['encounter_identifier'] = encounter_identifier
        status = encounter_resource.find("d:status", ns)
        if status != None:
            encounter["encounter_status"] = status.attrib["value"]
        _class = encounter_resource.find("d:class", ns)
        if _class != None:
            encounter["encounter_class"] = _class.find("d:code", ns).attrib["value"]
        _type = encounter_resource.find("d:type", ns)
        if _type != None:
            encounter["encounter_type"] = _type.find("d:text", ns).attrib["value"]
        service_type = encounter_resource.find("d:serviceType", ns)
        if service_type != None:
            encounter["encounter_service"] = service_type.find("d:text", ns).attrib[
                "value"
            ]
        # priority = encounter_resource.find('d:priority', ns)
        # if priority:
        #     encounter['encounter_priority'] = priority.find('d:text', ns).attrib['value']
        period = encounter_resource.find("d:period", ns)
        if period != None:
            encounter["encounter_start"] = period.find("d:start", ns).attrib["value"]
            end_date = None
            if period.find("d:end", ns) != None:
                end_date = period.find("d:end", ns).attrib["value"]
                encounter["encounter_end"] = getdatetime(end_date)
        length = encounter_resource.find("d:length", ns)
        if length != None:
            encounter["encounter_length"] = length.find("d:value", ns).attrib["value"]
        reason_code = encounter_resource.find("d:reasonCode", ns)
        if reason_code != None:
            encounter["encounter_reason"] = reason_code.find("d:text", ns).attrib[
                "value"
            ]
    return encounter


def get_service(service_resource, query_type):
    service = {}
    if query_type == "meta" or query_type == "all":
        meta = service_resource.find("d:meta", ns)
        service["version"] = meta.find("d:versionId", ns).attrib["value"]
        service["last_updated"] = meta.find("d:lastUpdated", ns).attrib["value"]
    if query_type == "id" or query_type == "all":
        service["id"] = service_resource.find("d:id", ns).attrib["value"]
    if query_type == "data" or query_type == "all":
        # service['service_identifier'] = service_identifier
        status = service_resource.find("d:status", ns)
        if status != None:
            service["service_status"] = status.attrib["value"]
        intent = service_resource.find("d:intent", ns)
        if intent != None:
            service["service_intent"] = intent.attrib["value"]
        category = service_resource.find("d:category", ns)
        if category != None:
            service["service_category"] = category.find("d:text", ns).attrib["value"]
        code = service_resource.find("d:code", ns)
        if code != None:
            service["service_code"] = code.find("d:text", ns).attrib["value"]
        occurrence = service_resource.find("d:occurrenceDateTime", ns)
        if occurrence != None:
            service["service_occurrence"] = get_date(occurrence.attrib["value"])
        authored_on = service_resource.find("d:authoredOn", ns)
        if authored_on != None:
            service["service_authored"] = get_date(authored_on.attrib["value"])
        note = service_resource.find("d:note", ns)
        if note:
            service["service_note"] = note.find("d:text", ns).attrib["value"]
    return service


def get_condition(condition_resource, query_type):
    condition = {}
    if query_type == "meta" or query_type == "all":
        meta = condition_resource.find("d:meta", ns)
        condition["version"] = meta.find("d:versionId", ns).attrib["value"]
        condition["last_updated"] = meta.find("d:lastUpdated", ns).attrib["value"]
    if query_type == "id" or query_type == "all":
        condition["id"] = condition_resource.find("d:id", ns).attrib["value"]
    if query_type == "data" or query_type == "all":
        # condition['condition_identifier'] = condition_identifier
        clinical_status = condition_resource.find("d:clinicalStatus", ns)
        if clinical_status != None:
            condition["condition_clinical_status"] = clinical_status.find(
                "d:text", ns
            ).attrib["value"]
        verification_status = condition_resource.find("d:verificationStatus", ns)
        if verification_status:
            condition["condition_verification_status"] = verification_status.find(
                "d:text", ns
            ).attrib["value"]
        category = condition_resource.find("d:category", ns)
        if category:
            condition["condition_category"] = category.find("d:text", ns).attrib[
                "value"
            ]
        severity = condition_resource.find("d:severity", ns)
        if severity:
            condition["condition_severity"] = severity.find("d:text", ns).attrib[
                "value"
            ]
        code = condition_resource.find("d:code", ns)
        if code:
            condition["condition_code"] = code.find("d:text", ns).attrib["value"]
        onset = condition_resource.find("d:onsetDateTime", ns)
        if onset != None:
            condition["condition_onset"] = onset.attrib["value"]
        abatement = condition_resource.find("d:abatementDateTime", ns)
        if abatement != None:
            condition["condition_abatement"] = abatement.attrib["value"]
        note = condition_resource.find("d:note", ns)
        if note:
            condition["condition_note"] = note.find("d:text", ns).attrib["value"]
    return condition


def get_observation(observation_resource, query_type):
    observation = {}
    if query_type == "meta" or query_type == "all":
        meta = observation_resource.find("d:meta", ns)
        observation["version"] = meta.find("d:versionId", ns).attrib["value"]
        observation["last_updated"] = meta.find("d:lastUpdated", ns).attrib["value"]
    if query_type == "id" or query_type == "all":
        observation["id"] = observation_resource.find("d:id", ns).attrib["value"]
    if query_type == "data" or query_type == "all":
        # observation['observation_identifier'] = observation_identifier
        status = observation_resource.find("d:status", ns)
        if status != None:
            observation["observation_status"] = status.attrib["value"]
        category = observation_resource.find("d:category", ns)
        if category:
            observation["observation_category"] = category.find("d:text", ns).attrib[
                "value"
            ]
        code = observation_resource.find("d:code", ns)
        if code:
            observation["observation_code"] = code.find("d:text", ns).attrib["value"]
        effective = observation_resource.find("d:effectiveDateTime", ns)
        if effective != None:
            observation["observation_effective"] = getdatetime(
                effective.attrib["value"]
            )
        value_quantity = observation_resource.find("d:valueQuantity", ns)
        if value_quantity:
            observation["observation_value_quantity"] = value_quantity.find(
                "d:value", ns
            ).attrib["value"]
            unit = value_quantity.find("d:unit", ns)
            if unit != None:
                observation["observation_value_unit"] = unit.attrib["value"]
        reference_range = observation_resource.find("d:referenceRange", ns)
        if reference_range:
            observation["observation_reference_range"] = reference_range.find(
                "d:text", ns
            ).attrib["value"]
    return observation


def get_procedure(procedure_resource, query_type):
    procedure = {}
    if query_type == "meta" or query_type == "all":
        meta = procedure_resource.find("d:meta", ns)
        procedure["version"] = meta.find("d:versionId", ns).attrib["value"]
        procedure["last_updated"] = meta.find("d:lastUpdated", ns).attrib["value"]
    if query_type == "id" or query_type == "all":
        procedure["id"] = procedure_resource.find("d:id", ns).attrib["value"]
    if query_type == "data" or query_type == "all":
        # procedure['procedure_identifier'] = procedure_identifier
        status = procedure_resource.find("d:status", ns)
        if status != None:
            procedure["procedure_status"] = status.attrib["value"]
        category = procedure_resource.find("d:category", ns)
        if category:
            procedure["procedure_category"] = category.find("d:text", ns).attrib[
                "value"
            ]
        code = procedure_resource.find("d:code", ns)
        if code:
            procedure["procedure_code"] = code.find("d:text", ns).attrib["value"]
        performed_datetime = procedure_resource.find("d:performedDateTime", ns)
        if performed_datetime != None:
            procedure["procedure_performed_datetime"] = getdatetime(
                performed_datetime.attrib["value"]
            )
        reason_code = procedure_resource.find("d:reasonCode", ns)
        if reason_code:
            procedure["procedure_reason_code"] = reason_code.find("d:text", ns).attrib[
                "value"
            ]
        outcome = procedure_resource.find("d:outcome", ns)
        if outcome:
            procedure["procedure_outcome"] = outcome.find("d:text", ns).attrib["value"]
        complication = procedure_resource.find("d:complication", ns)
        if complication:
            procedure["procedure_complication"] = complication.find(
                "d:text", ns
            ).attrib["value"]
        follow_up = procedure_resource.find("d:followUp", ns)
        if follow_up:
            procedure["procedure_follow_up"] = follow_up.find("d:text", ns).attrib[
                "value"
            ]
        note = procedure_resource.find("d:note", ns)
        if note:
            procedure["procedure_note"] = note.find("d:text", ns).attrib["value"]
    return procedure


def get_medication(medication_resource, query_type):
    medication_statement = {}
    if query_type == "meta" or query_type == "all":
        meta = medication_resource.find("d:meta", ns)
        medication_statement["version"] = meta.find("d:versionId", ns).attrib["value"]
        medication_statement["last_updated"] = meta.find("d:lastUpdated", ns).attrib[
            "value"
        ]
    if query_type == "id" or query_type == "all":
        medication_statement["id"] = medication_resource.find("d:id", ns).attrib[
            "value"
        ]
    if query_type == "data" or query_type == "all":
        # medication_statement['medication_identifier'] = medication_identifier
        medication = medication_resource.find("d:medicationCodeableConcept", ns)
        if medication:
            medication_statement["medication_medication"] = medication.find(
                "d:text", ns
            ).attrib["value"]
        effective = medication_resource.find("d:effectiveDateTime", ns)
        if effective != None:
            medication_statement["medication_effective"] = get_date(
                effective.attrib["value"]
            )
        date_asserted = medication_resource.find("d:dateAsserted", ns)
        if date_asserted != None:
            medication_statement["medication_date_asserted"] = get_date(
                date_asserted.attrib["value"]
            )
        reason_code = medication_resource.find("d:reasonCode", ns)
        if reason_code:
            medication_statement["medication_reason_code"] = reason_code.find(
                "d:text", ns
            ).attrib["value"]
        dosage = medication_resource.find("d:dosage", ns)
        if dosage:
            additional_instruction = dosage.find("d:additionalInstruction", ns)
            if additional_instruction:
                medication_statement[
                    "dosage_additional_instruction"
                ] = additional_instruction.find("d:text", ns).attrib["value"]
            patient_instruction = dosage.find("d:patientInstruction", ns)
            if patient_instruction:
                medication_statement[
                    "dosage_patient_instruction"
                ] = patient_instruction.find("d:text", ns).attrib["value"]
            timing = dosage.find("d:timing", ns)
            if timing:
                repeat = timing.find("d:repeat", ns)
                if repeat:
                    duration = repeat.find("d:duration", ns)
                    if duration != None:
                        duration_value = duration.attrib["value"]
                        duration_max = repeat.find("d:durationMax", ns)
                        if duration_max != None:
                            duration_value = (
                                duration.attrib["value"]
                                + "-"
                                + duration_max.attrib["value"]
                            )
                        duration_unit = repeat.find("d:durationUnit", ns).attrib[
                            "value"
                        ]
                        medication_statement["dosage_duration"] = duration_value
                        medication_statement["dosage_duration_unit"] = duration_unit
                    frequency = repeat.find("d:frequency", ns)
                    if frequency != None:
                        medication_statement["dosage_frequency"] = (
                            frequency.attrib["value"] + " lần"
                        )
                    period = repeat.find("d:period", ns)
                    if period != None:
                        period_value = period.attrib["value"]
                        period_max = repeat.find("d:periodMax", ns)
                        if period_max != None:
                            period_value = (
                                period.attrib["value"]
                                + "-"
                                + period_max.attrib["value"]
                            )
                        period_unit = repeat.find("d:periodUnit", ns)
                        medication_statement["dosage_period"] = period_value
                        medication_statement["dosage_period_value"] = period_unit
                    when = repeat.find("d:when", ns)
                    if when != None:
                        medication_statement["dosage_when"] = when.attrib["value"]
                    offset = repeat.find("d:offset", ns)
                    if offset != None:
                        medication_statement["dosage_offset"] = offset.attrib["value"]
            route = dosage.find("d:route", ns)
            if route:
                medication_statement["dosage_route"] = route.find("d:text", ns).attrib[
                    "value"
                ]
            dose_and_rate = dosage.find("d:doseAndRate", ns)
            if dose_and_rate:
                dose_quantity = dose_and_rate.find("d:doseQuantity", ns)
                quantity = dose_quantity.find("d:value", ns).attrib["value"]
                unit = dose_quantity.find("d:unit", ns).attrib["value"]
                medication_statement["dosage_quantity"] = quantity + " " + unit
    return medication_statement


def get_practitioner(practitioner_resource, query_type):
    practitioner = {}
    if query_type == "meta" or query_type == "all":
        meta = practitioner_resource.find("d:meta", ns)
        practitioner["version"] = meta.find("d:versionId", ns).attrib["value"]
        practitioner["last_updated"] = meta.find("d:lastUpdated", ns).attrib["value"]
    if query_type == "id" or query_type == "all":
        practitioner["id"] = practitioner_resource.find("d:id", ns).attrib["value"]
    if query_type == "data" or query_type == "all":
        # practitioner['identifier'] = practitioner_identifier
        practitioner_name = practitioner_resource.find("d:name", ns)
        if practitioner_name:
            practitioner["name"] = (
                practitioner_name.find("d:family", ns).attrib["value"]
                + " "
                + practitioner_name.find("d:given", ns).attrib["value"]
            )
        practitioner_telecom = practitioner_resource.find("d:telecom", ns)
        if practitioner_telecom:
            practitioner["telecom"] = practitioner_telecom.find("d:value", ns).attrib[
                "value"
            ]
        practitioner_gender = practitioner_resource.find("d:gender", ns)
        if practitioner_gender != None:
            if practitioner_gender.attrib["value"] == "male":
                practitioner["gender"] = "Nam"
            else:
                practitioner["gender"] = "Nữ"
        practitioner_birthdate = practitioner_resource.find("d:birthDate", ns)
        if practitioner_birthdate != None:
            practitioner["birthdate"] = practitioner_birthdate.attrib["value"]
        practitioner_qualification = practitioner_resource.find("d:qualification", ns)
        if practitioner_qualification:
            practitioner_qualification_code = practitioner_qualification.find(
                "d:code", ns
            )
            practitioner["qualification"] = practitioner_qualification_code.find(
                "d:text", ns
            ).attrib["value"]
    return practitioner


def get_diagnostic_report(diagnostic_report_resource, query_type):
    diagnostic_report = {}

    if query_type == "meta" or query_type == "all":
        meta = diagnostic_report_resource.find("d:meta", ns)
        diagnostic_report["version"] = meta.find("d:versionId", ns).attrib["value"]
        diagnostic_report["last_updated"] = getdatetime(
            meta.find("d:lastUpdated", ns).attrib["value"]
        )
    if query_type == "id" or query_type == "all":
        id_ = diagnostic_report_resource.find("d:id", ns)
        diagnostic_report["id"] = id_.attrib["value"]
    if query_type == "data" or query_type == "all":
        # diagnostic_report['diagnostic_identifier'] = diagnostic_report_identifier
        status = diagnostic_report_resource.find("d:status", ns)
        if status != None:
            diagnostic_report["diagnostic_status"] = status.attrib["value"]
        category = diagnostic_report_resource.find("d:category", ns)
        if category:
            diagnostic_report["diagnostic_category"] = category.find(
                "d:text", ns
            ).attrib["value"]

        code = diagnostic_report_resource.find("d:code", ns)
        if code:
            diagnostic_report["diagnostic_code"] = code.find("d:text", ns).attrib[
                "value"
            ]
        effective = diagnostic_report_resource.find("d:effectiveDateTime", ns)
        if effective != None:
            diagnostic_report["diagnostic_effective"] = effective.attrib["value"]
        conclusion = diagnostic_report_resource.find("d:conclusion", ns)
        if conclusion != None:
            diagnostic_report["diagnostic_conclusion"] = conclusion.attrib["value"]
    return diagnostic_report
