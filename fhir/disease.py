from os import error
import openpyxl as xl
import re
from fhir.models import DischargeDisease, ComorbidityDisease, Medicine, Test, Image


#Get discharge diseases:
wb = xl.load_workbook('fhir/ICD-10.xlsx', data_only=True)
sh = wb['ICD10']
max_row = sh.max_row
for i in range(4, max_row + 1):
    name_cell = sh.cell(row=i, column=20)
    code_cell = sh.cell(row=i, column=18)
    name = name_cell.value
    code = code_cell.value
    if len(code) == 4:
        code = list(code)
        code.insert(-1, '.')
        code = ''.join(code)
        print(code)
        DischargeDisease.objects.create(disease_code=code, disease_name=name, disease_search = name + ' ' + code)

#Get medicines:
wb = xl.load_workbook('fhir/Medicine.xlsx', data_only=True)
sh = wb['Medicine']
max_row = sh.max_row
for i in range(4, max_row + 1):
    medicine_name_cell = sh.cell(row=i, column=2)
    medicine_unit_cell = sh.cell(row=i, column=7)
    medicine_unit_price_cell = sh.cell(row=i, column=8)
    medicine_unit = medicine_unit_cell.value
    medicine_name = medicine_name_cell.value
    medicine_price = medicine_unit_price_cell.value
    modified_string = re.sub(r"\([^()]*\)", "", medicine_name)
    try:
        print('oke')
        Medicine.objects.create(medicine_name=modified_string.strip(), medicine_unit=medicine_unit, medicine_price_on_unit=medicine_price)
    except Exception as e:
        print(e)
    
#Get comorbidity diseases:
sh = wb['A1']
max_row = sh.max_row
diseases = {}
for i in range(4, max_row + 1):
    main_code = sh.cell(row=i, column=1)
    main_name = sh.cell(row=i, column=2)
    name_cell = sh.cell(row=i, column=6)
    if main_name.value != None:
        code_cell = sh.cell(row=i, column=3)
        name = name_cell.value
        code = code_cell.value
        if diseases.get(main_code.value):        
            diseases[main_code.value].append({'name': name, 'code': code})
        else:
            diseases[main_code.value] = []
            diseases[main_code.value].append({'name': name, 'code': code})
for key, value in diseases.items():
    try:
        main_disease = DischargeDisease.objects.get(disease_code=key)
        for disease in value:
            ComorbidityDisease.objects.create(discharge_diseases=main_disease, disease_code=disease['code'], disease_name=disease['name'], disease_search=disease['name'] + ' ' + disease['code'])
    except:
        pass


print('finished')

#Get tests

wb = xl.load_workbook('fhir/XETNGHIEM.xlsx', data_only=True)
sh = wb['Sheet1']
max_row = sh.max_row
for i in range(1, max_row + 1):
    test_name_cell = sh.cell(row=i, column=2)
    test_name = test_name_cell.value
    if test_name == None:
        category = sh.cell(row=i, column=1).value        
    try:
        Test.objects.create(test_name=test_name, test_category=category)
    except Exception as e:
        print(e)
        
        
#Get images
wb = xl.load_workbook('fhir/CHANDOANHINHANH.xlsx', data_only=True)
sh = wb['Sheet1']
max_row = sh.max_row
for i in range(1, max_row + 1):
    image_name_cell = sh.cell(row=i, column=2)
    image_name = image_name_cell.value
    if image_name == None:
        category = sh.cell(row=i, column=1).value
        continue
    try:
        print(image_name)
        Image.objects.create(image_name=image_name, image_category=category)
    except Exception as e:
        print(e)